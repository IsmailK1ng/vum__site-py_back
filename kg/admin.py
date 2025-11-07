from django.contrib import admin
from django.utils.html import format_html, escape
from django.http import HttpResponse
from django.utils.safestring import mark_safe
from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.conf import settings
from django.urls import path  
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reversion.admin import VersionAdmin

from .models import (
    KGVehicle, 
    KGVehicleImage, 
    VehicleCardSpec, 
    KGFeedback, 
    KGHeroSlide, 
    IconTemplate
)
from .forms import VehicleCardSpecForm

# ============================================
# БАЗОВЫЕ МИКСИНЫ ДЛЯ ПРАВ ДОСТУПА
# ============================================

class ContentAdminMixin:
    """Миксин для контент-админов KG"""
    def has_module_permission(self, request):
        if request.user.is_superuser:
            return True
        
        # Проверяем группы
        if request.user.groups.filter(
            name__in=['Главные админы', 'Контент KG', 'Контент UZ+KG']
        ).exists():
            return True
        
        # ✅ ПРОВЕРЯЕМ ИНДИВИДУАЛЬНЫЕ ПРАВА (любое право на просмотр контента KG)
        kg_models = ['kgvehicle', 'kgheroslide', 'icontemplate']
        for model in kg_models:
            if request.user.has_perm(f'kg.view_{model}'):
                return True
        
        return False
    
    def has_change_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        
        if request.user.groups.filter(
            name__in=['Главные админы', 'Контент KG', 'Контент UZ+KG']
        ).exists():
            return True
        
        # ✅ ПРОВЕРЯЕМ ИНДИВИДУАЛЬНОЕ ПРАВО
        model_name = self.model._meta.model_name
        return request.user.has_perm(f'kg.change_{model_name}')
    
    def has_delete_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        
        if request.user.groups.filter(name='Главные админы').exists():
            return True
        
        # ✅ ПРОВЕРЯЕМ ИНДИВИДУАЛЬНОЕ ПРАВО
        model_name = self.model._meta.model_name
        return request.user.has_perm(f'kg.delete_{model_name}')


class LeadManagerMixin:
    """Миксин для лид-менеджеров KG"""
    def has_module_permission(self, request):
        if request.user.is_superuser:
            return True
        
        # Проверяем группы
        if request.user.groups.filter(
            name__in=['Главные админы', 'Лиды KG', 'Лиды UZ+KG']
        ).exists():
            return True
        
        # ✅ ПРОВЕРЯЕМ ИНДИВИДУАЛЬНЫЕ ПРАВА
        return request.user.has_perm('kg.view_kgfeedback')
    
    def has_add_permission(self, request):
        return False  # Заявки создаются только с фронта
    
    def has_delete_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        
        if request.user.groups.filter(name='Главные админы').exists():
            return True
        
        # ✅ ПРОВЕРЯЕМ ИНДИВИДУАЛЬНОЕ ПРАВО
        return request.user.has_perm('kg.delete_kgfeedback')

class CustomReversionMixin:
    """Миксин для кастомного шаблона восстановления"""
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'recover/',
                self.admin_site.admin_view(self.custom_recover_list_view),
                name=f'{self.model._meta.app_label}_{self.model._meta.model_name}_recoverlist'
            ),
            path(
                'recover/<int:version_id>/',
                self.admin_site.admin_view(self.recover_view),
                name=f'{self.model._meta.app_label}_{self.model._meta.model_name}_recover'
            ),
        ]
        return custom_urls + urls
    
    def custom_recover_list_view(self, request):
        from reversion.models import Version
        from django.conf import settings
        
        opts = self.model._meta
        deleted_versions = Version.objects.get_deleted(self.model)
        
        seen_objects = {}
        version_list_with_preview = []
        
        for version in deleted_versions.order_by('-revision__date_created'):
            obj_repr = version.object_repr
            if obj_repr not in seen_objects:
                seen_objects[obj_repr] = True
                
                preview_url = None
                try:
                    field_dict = version.field_dict
                    for field_name in ['preview_image', 'main_image', 'card_image', 'logo']:
                        if field_name in field_dict and field_dict[field_name]:
                            preview_url = f"{settings.MEDIA_URL}{field_dict[field_name]}"
                            break
                except Exception:
                    pass
                
                version_list_with_preview.append({
                    'version': version,
                    'preview_url': preview_url,
                    'object_repr': obj_repr,
                    'date': version.revision.date_created
                })
        
        context = {
            **self.admin_site.each_context(request),
            'opts': opts,
            'version_list': version_list_with_preview,
            'title': f'Восстановление: {opts.verbose_name_plural}',
            'has_view_permission': self.has_view_permission(request),
        }
        
        return render(request, 'admin/reversion/recover_list.html', context)
    
    def recover_view(self, request, version_id):
        from reversion.models import Version
        from django.contrib import messages
        from django.shortcuts import redirect
        
        opts = self.model._meta
        
        try:
            version = Version.objects.get(pk=version_id)
            
            if version.content_type.model_class() != self.model:
                raise Version.DoesNotExist
            
            version.revision.revert()
            
            messages.success(request, f'✅ Объект "{version.object_repr}" успешно восстановлен!')
            return redirect(f'admin:{opts.app_label}_{opts.model_name}_changelist')
            
        except Version.DoesNotExist:
            messages.error(request, '❌ Версия не найдена или уже восстановлена')
            return redirect(f'admin:{opts.app_label}_{opts.model_name}_recoverlist')
        except Exception as e:
            messages.error(request, f'❌ Ошибка восстановления: {str(e)}')
            return redirect(f'admin:{opts.app_label}_{opts.model_name}_recoverlist')

# ============================================
# INLINE: ХАРАКТЕРИСТИКИ ДЛЯ КАТАЛОГА
# ============================================

class VehicleCardSpecInline(admin.TabularInline):
    model = VehicleCardSpec
    form = VehicleCardSpecForm 
    extra = 1
    fields = ('icon_selector', 'icon_preview', 'value_ru', 'order')
    readonly_fields = ('icon_selector', 'icon_preview')
    verbose_name = "Характеристика"
    verbose_name_plural = "Характеристики для каталога (отображаются в карточках на главной странице)"

    class Media:
        js = ('admin/js/icon_selector.js',)

    def icon_preview(self, obj):
        if obj and obj.icon:
            return format_html(
                '<img src="{}" width="50" style="border-radius:8px;" alt="Иконка">',
                obj.icon.url
            )
        return "—"
    icon_preview.short_description = "Превью"

    def icon_selector(self, obj):
        templates = IconTemplate.objects.all().order_by('order')
        
        if not templates.exists():
            return mark_safe('<p style="color:red;">Нет иконок в библиотеке</p>')
        
        html_parts = [
            '<div class="icon-selector-widget">',
            '<details style="border:1px solid #ddd; border-radius:8px; padding:8px; background:#f9f9f9;">',
            f'<summary style="cursor:pointer; padding:10px; background:#e3f2fd; border-radius:6px; font-weight:600;">Выбрать иконку ({templates.count()})</summary>',
            '<div style="padding:10px 0; margin-top:10px;">',
            '<div class="icon-grid" style="display:grid; grid-template-columns:repeat(2, 1fr); gap:10px; max-width:220px;">'
        ]
        
        for t in templates:
            html_parts.append(
                f'<div class="icon-card" data-template-id="{t.id}" data-icon-url="{t.icon.url}" '
                f'style="text-align:center; padding:8px; border:2px solid #ddd; border-radius:8px; '
                f'cursor:pointer; background:#fff; transition:all 0.2s;">'
                f'<img src="{t.icon.url}" width="40" height="40" style="display:block; margin:0 auto 5px;" alt="{escape(t.name)}">'
                f'<small style="font-size:9px; color:#666;">{escape(t.name)}</small>'
                f'</div>'
            )
        
        html_parts.extend([
            '</div></div></details></div>',
            '<style>.icon-card:hover{border-color:#64b5f6!important;transform:scale(1.05);}</style>'
        ])
        
        return mark_safe(''.join(html_parts))
    icon_selector.short_description = "Выбор иконки"


# ============================================
# INLINE: ДОПОЛНИТЕЛЬНЫЕ ФОТО
# ============================================

class KGVehicleImageInline(admin.TabularInline):
    model = KGVehicleImage
    extra = 1
    fields = ('image', 'image_preview', 'alt', 'order')
    readonly_fields = ('image_preview',)
    verbose_name = "Дополнительное фото"
    verbose_name_plural = "Дополнительные фото (для галереи на странице детализации)"

    def image_preview(self, obj):
        if obj and obj.image:
            return format_html(
                '<img src="{}" width="80" style="border-radius:8px;" alt="Превью">',
                obj.image.url
            )
        return "Не загружено"
    image_preview.short_description = "Превью"


# ============================================
# ADMIN: КАТАЛОГ МАШИН
# ============================================

@admin.register(KGVehicle)
class KGVehicleAdmin(ContentAdminMixin, CustomReversionMixin, VersionAdmin, admin.ModelAdmin):  
    history_latest_first = True 
    list_display = ('mini_thumb', 'title_display', 'category_badge', 'is_active', 'created_at', 'action_buttons')
    list_editable = ('is_active',)
    list_filter = ('category', 'is_active', 'created_at')
    list_select_related = True
    search_fields = ('title_ru', 'title_ky', 'title_en', 'slug')
    readonly_fields = ('created_at', 'updated_at', 'category')
    list_per_page = 20
    date_hierarchy = 'created_at'
    inlines = [VehicleCardSpecInline, KGVehicleImageInline]

    fieldsets = (
        ('Название техники', {
            'fields': ('title_ru', 'title_ky', 'title_en'),
            'description': 'Это название отображается в каталоге, в Hero-слайдере и на странице детализации машины.'
        }),
        ('Фотографии', {
            'fields': ('preview_image', 'main_image'),
            'description': 'Превью — для каталога (главная страница). Главное фото — для страницы детализации.'
        }),
        ('Детальные характеристики', {
            'fields': (
                'feature_aircondi', 'feature_power_windows', 'feature_sleeping_area',
                'feature_radio', 'feature_remote_control', 'feature_bluetooth',
                'feature_multifunction_steering',
                
                'wheel_formula', 'dimensions_ru', 'wheelbase',
                'fuel_type_ru', 'fuel_type_ky', 'fuel_type_en', 'tank_volume',
                
                'curb_weight', 'payload', 'gross_weight',
                
                'body_type_ru', 'body_type_ky', 'body_type_en',
                'body_dimensions_ru', 'body_volume',
                'body_material_ru', 'body_material_ky', 'body_material_en',
                'loading_type_ru', 'loading_type_ky', 'loading_type_en',
                
                'engine_model', 'engine_volume', 'engine_power',
                
                'transmission_model', 'transmission_type_ru', 'transmission_type_ky',
                'transmission_type_en', 'gears',
                
                'tire_type', 'suspension_ru', 'suspension_ky', 'suspension_en',
                'brakes_ru', 'brakes_ky', 'brakes_en',
                
                'cabin_category_ru', 'cabin_category_ky', 'cabin_category_en',
                'cabin_equipment_ru', 'cabin_equipment_ky', 'cabin_equipment_en',
            ),
            'classes': ('specs-container',),
            'description': 'Эта вкладка содержит подробные технические характеристики машины, которые отображаются на странице детализации. Используйте аккордеоны для удобной навигации.'
        }),
        ('Служебная информация', {
            'fields': ('is_active', 'category', 'created_at', 'updated_at'),
            'classes': ('collapse',),
            'description': 'Служебные поля для управления видимостью и отслеживания изменений.'
        }),
    )

    class Media:
        css = {'all': ('admin/css/vehicle_admin.css',)}
        js = ('admin/js/specs_accordion.js',)

    def get_queryset(self, request):
        return super().get_queryset(request).select_related().prefetch_related('card_specs', 'mini_images')

    def mini_thumb(self, obj):
        if obj.preview_image:
            return format_html(
                '<img src="{}" width="60" style="border-radius:6px;" alt="Фото">',
                obj.preview_image.url
            )
        return "—"
    mini_thumb.short_description = "Фото"

    def title_display(self, obj):
        return obj.title_ru or obj.title or '—'
    title_display.short_description = "Название"
    title_display.admin_order_field = 'title_ru'

    def category_badge(self, obj):
        colors = {'v': '#4CAF50', 'vr': '#2196F3', 'vh': '#FF9800'}
        labels = {'v': 'V Series', 'vr': 'VR Series', 'vh': 'VH Series'}
        return format_html(
            '<span style="background:{}; color:white; padding:5px 12px; border-radius:12px; font-weight:600; font-size:11px;">{}</span>',
            colors.get(obj.category, '#757575'), 
            labels.get(obj.category, obj.category.upper())
        )
    category_badge.short_description = "Серия"

    def action_buttons(self, obj):
        frontend_url = getattr(settings, 'FRONTEND_URL', 'http://localhost:3000')
        vehicle_url = f"{frontend_url}/vehicle-details.html?id={obj.slug_ru or obj.slug}&lang=ru"
        title_safe = escape(obj.title_ru or 'машину')
        
        return format_html(
            '<div style="display:flex; gap:8px;">'
            '<a href="{}" title="Редактировать"><img src="/static/media/icon-adminpanel/pencil.png" width="28" alt="Редактировать"></a>'
            '<a href="{}" onclick="return confirm(\'Удалить {}?\')" title="Удалить"><img src="/static/media/icon-adminpanel/recycle-bin.png" width="28" alt="Удалить"></a>'
            '<a href="{}" target="_blank" rel="noopener noreferrer" title="Посмотреть"><img src="/static/media/icon-adminpanel/eyes.png" width="28" alt="Посмотреть"></a>'
            '</div>',
            f'/admin/kg/kgvehicle/{obj.id}/change/',
            f'/admin/kg/kgvehicle/{obj.id}/delete/',
            title_safe,
            vehicle_url
        )
    action_buttons.short_description = "Действия"

    actions = ['activate_vehicles', 'deactivate_vehicles']

    def activate_vehicles(self, request, queryset):
        updated = queryset.update(is_active=True)
        self.message_user(request, f'Активировано: {updated}')
    activate_vehicles.short_description = 'Активировать'

    def deactivate_vehicles(self, request, queryset):
        updated = queryset.update(is_active=False)
        self.message_user(request, f'Деактивировано: {updated}')
    deactivate_vehicles.short_description = 'Деактивировать'

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        from reversion.models import Version
        deleted_count = Version.objects.get_deleted(self.model).count()
        if deleted_count > 0:
            extra_context['show_recover_button'] = True
            extra_context['deleted_count'] = deleted_count
        return super().changelist_view(request, extra_context)


# ============================================
# ADMIN: HERO-СЛАЙДЫ
# ============================================

@admin.register(KGHeroSlide)
class KGHeroSlideAdmin(ContentAdminMixin, CustomReversionMixin, VersionAdmin, admin.ModelAdmin):  
    history_latest_first = True 
    list_display = ('order', 'vehicle_display', 'is_active', 'created_at')
    list_display_links = ('vehicle_display',)
    list_editable = ('is_active', 'order')
    list_filter = ('is_active', 'created_at')
    search_fields = ('vehicle__title_ru', 'description_ru')
    readonly_fields = ('created_at', 'vehicle_preview')
    list_select_related = ('vehicle',)

    fieldsets = (
        ('Основная информация', {
            'fields': ('vehicle', 'vehicle_preview', 'order', 'is_active'),
            'description': ' Hero-слайды отображаются на главной странице сайта. Выберите машину и настройте порядок показа.'
        }),
        ('Описание (Русский)', {'fields': ('description_ru',), 'description': 'Краткое описание машины для главной страницы (например: "Мощный и надежный грузовик").'}),
        ('Описание (Кыргызский)', {'fields': ('description_ky',)}),
        ('Описание (Английский)', {'fields': ('description_en',)}),
    )

    def vehicle_display(self, obj):
        return obj.vehicle.title_ru if obj.vehicle else '—'
    vehicle_display.short_description = "Машина"
    vehicle_display.admin_order_field = 'vehicle__title_ru'

    def vehicle_preview(self, obj):
        if not obj.vehicle:
            return "—"
        img = obj.vehicle.main_image or obj.vehicle.preview_image
        if img:
            return format_html(
                '<img src="{}" width="200" style="border-radius:8px;" alt="Превью машины">',
                img.url
            )
        return "—"
    vehicle_preview.short_description = "Превью"

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        from reversion.models import Version
        deleted_count = Version.objects.get_deleted(self.model).count()
        if deleted_count > 0:
            extra_context['show_recover_button'] = True
            extra_context['deleted_count'] = deleted_count
        return super().changelist_view(request, extra_context)


# ============================================
# ADMIN: ЗАЯВКИ
# ============================================

@admin.register(KGFeedback)
class KGFeedbackAdmin(LeadManagerMixin, admin.ModelAdmin):  
    list_display = ['name', 'phone', 'region', 'vehicle_display', 'priority', 'status', 'manager', 'created_at', 'action_buttons']
    list_editable = ['priority', 'status', 'manager']
    list_filter = ['status', 'priority', 'region', 'created_at']
    search_fields = ['name', 'phone', 'vehicle__title_ru']
    readonly_fields = ['created_at']
    date_hierarchy = 'created_at'
    actions = ['export_to_excel', 'mark_as_done']
    list_select_related = ('vehicle', 'manager')
    raw_id_fields = ['vehicle', 'manager']

    fieldsets = (
        ('Информация о клиенте', {
            'fields': ('name', 'phone', 'region', 'vehicle', 'message', 'created_at'),
        }),
        ('Управление', {
            'fields': ('status', 'priority', 'manager', 'admin_comment'),
        }),
    )

    class Media:
        css = {'all': ('admin/css/feedback_admin.css',)}
        js = ('admin/js/auto_save_feedback.js',)

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        if request.user.is_superuser or request.user.groups.filter(name='Главные админы').exists():
            extra_context['stats_button_html'] = format_html(
                '<div style="margin-bottom: 20px;">'
                '<a href="/admin/kg/stats/" target="_blank" style="'
                'background: linear-gradient(306deg, #000000, #002b9b, #1e57eb); '
                'color: white; padding: 12px 24px; border: none; border-radius: 8px; '
                'font-size: 15px; font-weight: 600; text-decoration: none; '
                'display: inline-block; transition: all 0.3s; '
                'box-shadow: 0 4px 12px rgba(0, 43, 155, 0.2);" '
                'onmouseover="this.style.transform=\'translateY(-2px)\'; this.style.boxShadow=\'0 6px 20px rgba(0, 43, 155, 0.3)\';" '
                'onmouseout="this.style.transform=\'translateY(0)\'; this.style.boxShadow=\'0 4px 12px rgba(0, 43, 155, 0.2)\';">'
                ' Статистика заявок</a></div>'
            )
        return super().changelist_view(request, extra_context)

    def vehicle_display(self, obj):
        return obj.vehicle.title_ru if obj.vehicle else '—'
    vehicle_display.short_description = "Машина"
    vehicle_display.admin_order_field = 'vehicle__title_ru'

    def action_buttons(self, obj):
        frontend_url = getattr(settings, 'FRONTEND_URL', 'http://localhost:3000')
        vehicle_url = f"{frontend_url}/vehicle-details.html?id={obj.vehicle.slug_ru}&lang=ru" if obj.vehicle else "#"
        name_safe = escape(obj.name)
        
        return format_html(
            '<div class="action-buttons">'
            '<a href="{}" title="Редактировать"><img src="/static/media/icon-adminpanel/pencil.png" width="28" alt="Редактировать"></a>'
            '<a href="{}" onclick="return confirm(\'Удалить заявку от {}?\')" title="Удалить"><img src="/static/media/icon-adminpanel/recycle-bin.png" width="28" alt="Удалить"></a>'
            '<a href="{}" target="_blank" rel="noopener noreferrer" title="Посмотреть машину"><img src="/static/media/icon-adminpanel/eyes.png" width="28" alt="Посмотреть"></a>'
            '</div>',
            f'/admin/kg/kgfeedback/{obj.id}/change/',
            f'/admin/kg/kgfeedback/{obj.id}/delete/',
            name_safe,
            vehicle_url
        )
    action_buttons.short_description = "Действия"

    def mark_as_done(self, request, queryset):
        updated = queryset.update(status='done')
        self.message_user(request, f'Обработано заявок: {updated}')
    mark_as_done.short_description = 'Отметить как обработанные'

    def export_to_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Заявки FAW KG"

        headers = ['№', 'ФИО', 'Телефон', 'Регион', 'Машина', 'Статус', 'Приоритет', 'Менеджер', 'Дата']
        ws.append(headers)

        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Оптимизация: создаём словари один раз
        REGION_DICT = dict(KGFeedback.REGION_CHOICES)
        STATUS_DICT = dict(KGFeedback.STATUS_CHOICES)
        PRIORITY_DICT = dict(KGFeedback.PRIORITY_CHOICES)

        for idx, feedback in enumerate(queryset.select_related('vehicle', 'manager'), start=1):
            ws.append([
                idx,
                feedback.name,
                feedback.phone,
                REGION_DICT.get(feedback.region, feedback.region),
                feedback.vehicle.title_ru if feedback.vehicle else '-',
                STATUS_DICT.get(feedback.status, feedback.status),
                PRIORITY_DICT.get(feedback.priority, feedback.priority),
                feedback.manager.username if feedback.manager else '-',
                feedback.created_at.strftime('%d.%m.%Y %H:%M')
            ])

        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="faw_kg_leads_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response

    export_to_excel.short_description = 'Экспорт в Excel'


# ============================================
# ADMIN: ШАБЛОНЫ ИКОНОК
# ============================================

@admin.register(IconTemplate)
class IconTemplateAdmin(ContentAdminMixin, admin.ModelAdmin):  
    list_display = ('icon_preview', 'name', 'order')  
    list_editable = ('order',)
    fields = ('name', 'icon', 'icon_preview', 'order')
    readonly_fields = ('icon_preview',)
    list_per_page = 50
    
    def icon_preview(self, obj):
        if obj and obj.icon:
            return format_html(
                '<img src="{}" width="60" style="border-radius:8px;" alt="{}">',
                obj.icon.url,
                escape(obj.name)
            )
        return "—"
    icon_preview.short_description = "Превью"
