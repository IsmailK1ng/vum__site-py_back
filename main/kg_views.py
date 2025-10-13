# main/kg_views.py
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAdminUser
from django.db.models import Count
from datetime import datetime, timedelta
from django.http import HttpResponse
from .models import KGVehicle, KGFeedback, KGHeroSlide
from .kg_serializers import (
    KGVehicleSerializer, 
    KGFeedbackSerializer,
    KGHeroSlideSerializer
)


class KGVehicleViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API для каталога машин faw.kg
    - list: каталог для главной страницы
    - retrieve: детальная информация для vehicle-details.ts
    """
    queryset = KGVehicle.objects.filter(is_active=True).prefetch_related(
        'vehicle_features__feature', 'mini_images'
    )
    permission_classes = [AllowAny]
    lookup_field = 'slug'
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return KGVehicleDetailSerializer
        return KGVehicleListSerializer


class KGFeedbackViewSet(viewsets.ModelViewSet):
    """
    API для заявок с сайта faw.kg
    """
    queryset = KGFeedback.objects.all().select_related('vehicle').order_by('-created_at')
    
    def get_serializer_class(self):
        if self.action == 'create':
            return KGFeedbackCreateSerializer
        return KGFeedbackSerializer
    
    def get_permissions(self):
        if self.action == 'create':
            return [AllowAny()]
        return [IsAdminUser()]
    
    def create(self, request):
        """Создание новой заявки"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        feedback = serializer.save()
        
        return Response(
            {
                'success': True,
                'message': 'Ваша заявка принята. Мы свяжемся с вами в ближайшее время.',
                'data': KGFeedbackSerializer(feedback).data
            },
            status=status.HTTP_201_CREATED
        )
    
    @action(detail=False, methods=['get'], permission_classes=[IsAdminUser])
    def export_excel(self, request):
        """
        Экспорт лидов в Excel
        Параметры: ?start_date=2025-01-01&end_date=2025-12-31&region=Bishkek&vehicle_id=1
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response(
                {'error': 'openpyxl не установлен. Выполните: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        # Фильтры
        queryset = self.get_queryset()
        
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        region = request.query_params.get('region')
        vehicle_id = request.query_params.get('vehicle_id')
        is_processed = request.query_params.get('is_processed')
        
        if start_date:
            queryset = queryset.filter(created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__lte=end_date)
        if region:
            queryset = queryset.filter(region=region)
        if vehicle_id:
            queryset = queryset.filter(vehicle_id=vehicle_id)
        if is_processed is not None:
            queryset = queryset.filter(is_processed=is_processed.lower() == 'true')
        
        # Создание Excel файла
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Заявки FAW KG"
        
        # Заголовки
        headers = ['№', 'ФИО', 'Телефон', 'Регион', 'Машина', 'Сообщение', 'Дата', 'Обработано', 'Комментарий']
        ws.append(headers)
        
        # Стилизация заголовков
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Данные
        for idx, feedback in enumerate(queryset, start=1):
            ws.append([
                idx,
                feedback.name,
                feedback.phone,
                feedback.get_region_display(),
                feedback.vehicle.title if feedback.vehicle else '-',
                feedback.message or '-',
                feedback.created_at.strftime('%d.%m.%Y %H:%M'),
                'Да' if feedback.is_processed else 'Нет',
                feedback.admin_comment or '-'
            ])
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Отправка файла
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="faw_kg_leads_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        
        return response
    
    @action(detail=False, methods=['get'], permission_classes=[IsAdminUser])
    def statistics(self, request):
        """Статистика по заявкам"""
        total = self.queryset.count()
        processed = self.queryset.filter(is_processed=True).count()
        unprocessed = self.queryset.filter(is_processed=False).count()
        today = self.queryset.filter(created_at__date=datetime.now().date()).count()
        this_week = self.queryset.filter(
            created_at__gte=datetime.now() - timedelta(days=7)
        ).count()
        this_month = self.queryset.filter(
            created_at__month=datetime.now().month,
            created_at__year=datetime.now().year
        ).count()
        
        by_region = self.queryset.values('region').annotate(
            count=Count('id')
        ).order_by('-count')
        
        by_vehicle = self.queryset.filter(vehicle__isnull=False).values(
            'vehicle__title'
        ).annotate(count=Count('id')).order_by('-count')[:10]
        
        return Response({
            'total': total,
            'processed': processed,
            'unprocessed': unprocessed,
            'today': today,
            'this_week': this_week,
            'this_month': this_month,
            'by_region': list(by_region),
            'top_vehicles': list(by_vehicle)
        })
    
    @action(detail=True, methods=['post'], permission_classes=[IsAdminUser])
    def mark_processed(self, request, pk=None):
        """Отметить заявку как обработанную"""
        feedback = self.get_object()
        feedback.is_processed = True
        feedback.admin_comment = request.data.get('comment', feedback.admin_comment)
        feedback.save()
        
        return Response({
            'status': 'processed',
            'message': f'Заявка #{feedback.id} отмечена как обработанная'
        })




class KGHeroSlideViewSet(viewsets.ReadOnlyModelViewSet):
    """API для Hero-слайдов на главной странице"""
    queryset = KGHeroSlide.objects.filter(is_active=True).select_related('vehicle')
    serializer_class = KGHeroSlideSerializer
    permission_classes = [AllowAny]