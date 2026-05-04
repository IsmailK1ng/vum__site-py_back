# ========== PYTHON STANDARD LIBRARY ==========
import asyncio
import json
import logging
import openpyxl
import os
import subprocess
from datetime import datetime, timedelta
from urllib.parse import unquote

# ========== DJANGO ==========
from django import forms
from django.conf import settings
from django.contrib import admin, messages
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Q, Max, Count, Case, When, IntegerField
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect
from django.urls import path
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from django.views.decorators.cache import never_cache

# ========== DJANGO THIRD-PARTY ==========
from modeltranslation.admin import (
    TranslationTabularInline,
    TranslationStackedInline,
    TabbedTranslationAdmin,
)
from reversion.admin import VersionAdmin
from reversion.models import Version
from openpyxl.styles import Font, PatternFill, Alignment

# ========== AIOGRAM ==========
from aiogram import Bot
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.exceptions import TelegramForbiddenError, TelegramBadRequest

# ========== ЛОКАЛЬНЫЕ ИМПОРТЫ ==========
from .models import (
    News, NewsBlock, ContactForm, Vacancy,
    VacancyResponsibility, VacancyRequirement,
    VacancyCondition, VacancyIdealCandidate,
    JobApplication, FeatureIcon, Product,
    ProductParameter, ProductFeature, ProductCardSpec,
    ProductGallery, DealerService, Dealer,
    BecomeADealerPage, DealerRequirement,
    BecomeADealerApplication, AmoCRMToken,
    Dashboard, Promotion, PageMeta, FAQItem,
    REGION_CHOICES,
    TelegramUser, TestDriveRequest, BotConfig,
    BotMessage, BotMenuItem, BotBroadcast,
    ProductWishlist, ProductViewHistory, BotContacts,
    TeamDepartment, TeamMember,
)
from .forms import PageMetaAdminForm
from main.services.amocrm.token_manager import TokenManager
from main.services.amocrm.lead_sender import LeadSender
from main.services.dashboard.analytics import calculate_kpi
from main.services.dashboard.charts import get_chart_data
from main.services.dashboard.insights import generate_insights

logger = logging.getLogger('bot')

# ========== НАСТРОЙКИ АДМИНКИ ==========
admin.site.site_header = "Панель управления VUM"
admin.site.site_title = "VUM Admin"
admin.site.index_title = "Управление сайтами FAW"

BOT_SERVICE_NAME = 'faw_bot'


# ============ БАЗОВЫЕ МИКСИНЫ ============

class ContentAdminMixin:
    """Миксин для контент-админов"""

    def has_module_permission(self, request):
        if request.user.is_superuser:
            return True
        if request.user.groups.filter(
            name__in=['Главные админы', 'Контент-админы', 'Контент UZ', 'Контент UZ+KG']
        ).exists():
            return True
        content_models = [
            'news', 'product', 'vacancy', 'dealer', 'dealerservice',
            'featureicon', 'becomeadealerpage',
        ]
        return any(request.user.has_perm(f'main.view_{m}') for m in content_models)

    def has_change_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        if request.user.groups.filter(
            name__in=['Главные админы', 'Контент-админы', 'Контент UZ', 'Контент UZ+KG']
        ).exists():
            return True
        model_name = self.model._meta.model_name
        return request.user.has_perm(f'main.change_{model_name}')

    def has_delete_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        if request.user.groups.filter(name='Главные админы').exists():
            return True
        model_name = self.model._meta.model_name
        return request.user.has_perm(f'main.delete_{model_name}')


class LeadManagerMixin:
    """Миксин для лид-менеджеров"""

    def has_module_permission(self, request):
        if request.user.is_superuser:
            return True
        if request.user.groups.filter(
            name__in=['Главные админы', 'Лид-менеджеры', 'Лиды UZ', 'Лиды UZ+KG']
        ).exists():
            return True
        lead_models = ['contactform', 'jobapplication', 'becomeadealerapplication']
        return any(request.user.has_perm(f'main.view_{m}') for m in lead_models)

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        if request.user.groups.filter(name='Главные админы').exists():
            return True
        model_name = self.model._meta.model_name
        return request.user.has_perm(f'main.delete_{model_name}')


class AmoCRMAdminMixin:
    """Миксин для управления amoCRM"""

    def has_module_permission(self, request):
        if request.user.is_superuser:
            return True
        return request.user.groups.filter(name='Главные админы').exists()

    def has_view_permission(self, request, obj=None):
        return self.has_module_permission(request)

    def has_change_permission(self, request, obj=None):
        return self.has_module_permission(request)

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


class CustomReversionMixin:
    """Миксин для кастомного шаблона восстановления"""

    def get_urls(self):
        urls = super().get_urls()
        opts = self.model._meta
        custom_urls = [
            path(
                'recover/',
                self.admin_site.admin_view(self.custom_recover_list_view),
                name=f'{opts.app_label}_{opts.model_name}_recoverlist',
            ),
            path(
                'recover/<int:version_id>/',
                self.admin_site.admin_view(self.recover_view),
                name=f'{opts.app_label}_{opts.model_name}_recover',
            ),
        ]
        return custom_urls + urls

    def custom_recover_list_view(self, request):
        opts = self.model._meta
        deleted_versions = Version.objects.get_deleted(self.model)
        seen_objects = {}
        version_list_with_preview = []

        for version in deleted_versions.order_by('-revision__date_created'):
            obj_repr = version.object_repr
            if obj_repr in seen_objects:
                continue
            seen_objects[obj_repr] = True
            preview_url = None
            try:
                field_dict = version.field_dict
                for field_name in ['preview_image', 'main_image', 'card_image', 'logo']:
                    if field_name in field_dict and field_dict[field_name]:
                        preview_url = f"{settings.MEDIA_URL}{field_dict[field_name]}"
                        break
            except Exception:
                pass
            version_list_with_preview.append({
                'version': version,
                'preview_url': preview_url,
                'object_repr': obj_repr,
                'date': version.revision.date_created,
            })

        context = {
            **self.admin_site.each_context(request),
            'opts': opts,
            'version_list': version_list_with_preview,
            'title': f'Восстановление: {opts.verbose_name_plural}',
            'has_view_permission': self.has_view_permission(request),
        }
        return render(request, 'admin/reversion/recover_list.html', context)

    def recover_view(self, request, version_id):
        opts = self.model._meta
        try:
            version = Version.objects.get(pk=version_id)
            if version.content_type.model_class() != self.model:
                raise Version.DoesNotExist
            version.revision.revert()
            messages.success(request, f'Объект "{version.object_repr}" успешно восстановлен!')
            return redirect(f'admin:{opts.app_label}_{opts.model_name}_changelist')
        except Version.DoesNotExist:
            messages.error(request, 'Версия не найдена или уже восстановлена')
            return redirect(f'admin:{opts.app_label}_{opts.model_name}_recoverlist')
        except Exception as exc:
            messages.error(request, f'Ошибка восстановления: {exc}')
            return redirect(f'admin:{opts.app_label}_{opts.model_name}_recoverlist')


def _add_recover_context(admin_instance, request, extra_context):
    """Хелпер — добавляет кнопку восстановления в changelist если есть удалённые."""
    extra_context = extra_context or {}
    deleted_count = Version.objects.get_deleted(admin_instance.model).count()
    if deleted_count > 0:
        extra_context['show_recover_button'] = True
        extra_context['deleted_count'] = deleted_count
    return extra_context


# ============ НОВОСТИ ============

class NewsBlockInline(TranslationStackedInline):
    model = NewsBlock
    extra = 1
    fields = ('block_type', 'title', 'text', 'image', 'youtube_url', 'video_file', 'order')

    class Media:
        js = ('js/admin/news_block_dynamic.js',)
        css = {'all': ('css/news_block_custom.css',)}


@admin.register(News)
class NewsAdmin(ContentAdminMixin, CustomReversionMixin, VersionAdmin, TabbedTranslationAdmin):
    list_display = ['preview_image_tag', 'title', 'author', 'is_active', 'order', 'created_at', 'seo_button', 'action_buttons']
    list_editable = ['is_active', 'order']
    list_filter = ['is_active', 'created_at']
    search_fields = ['title', 'desc']
    readonly_fields = ['preview_image_tag', 'author_photo_tag', 'slug', 'updated_at']
    inlines = [NewsBlockInline]
    history_latest_first = True

    fieldsets = (
        ('Основная информация', {
            'fields': ('title', 'slug', 'created_at', 'is_active', 'order'),
        }),
        ('Карточка новости', {
            'fields': ('desc', 'preview_image', 'preview_image_tag'),
        }),
        ('Автор', {
            'fields': ('author', 'author_photo', 'author_photo_tag'),
        }),
        ('Техническая информация', {
            'fields': ('updated_at',),
            'classes': ('collapse',),
        }),
    )

    def preview_image_tag(self, obj):
        if obj.preview_image:
            return format_html('<img src="{}" width="100" style="border-radius:8px;"/>', obj.preview_image.url)
        return '—'
    preview_image_tag.short_description = 'Превью'

    def author_photo_tag(self, obj):
        if obj.author_photo:
            return format_html('<img src="{}" width="50" style="border-radius:50%;">', obj.author_photo.url)
        return '—'
    author_photo_tag.short_description = 'Фото автора'

    def action_buttons(self, obj):
        return format_html(
            '<div style="display:flex;gap:8px;">'
            '<a href="{}"><img src="/static/media/icon-adminpanel/pencil.png" width="24" height="24"></a>'
            '<a href="/news/{}/" target="_blank"><img src="/static/media/icon-adminpanel/eyes.png" width="24" height="24"></a>'
            '<a href="{}" onclick="return confirm(\'Удалить?\')"><img src="/static/media/icon-adminpanel/recycle-bin.png" width="24" height="24"></a>'
            '</div>',
            f'/admin/main/news/{obj.id}/change/',
            obj.slug,
            f'/admin/main/news/{obj.id}/delete/',
        )
    action_buttons.short_description = 'Действия'

    def seo_button(self, obj):
        try:
            seo = PageMeta.objects.get(model='Post', key=str(obj.id))
            color = '#10b981' if seo.is_active else '#ef4444'
            status = 'Активно' if seo.is_active else 'Неактивно'
            return format_html(
                '<a href="/admin/main/pagemeta/{}/change/" title="{}" '
                'style="background:{};color:white;padding:6px 12px;border-radius:6px;'
                'text-decoration:none;font-weight:600;font-size:11px;">SEO</a>',
                seo.id, status, color,
            )
        except PageMeta.DoesNotExist:
            return format_html('<span style="color:#999;font-size:11px;">—</span>')
    seo_button.short_description = 'SEO'

    def changelist_view(self, request, extra_context=None):
        return super().changelist_view(request, _add_recover_context(self, request, extra_context))


# ============ ЗАЯВКИ ============

@admin.register(ContactForm)
class ContactFormAdmin(LeadManagerMixin, admin.ModelAdmin):
    change_list_template = 'main/contactform/change_list.html'
    preserve_filters = True
    list_select_related = ['manager']
    list_per_page = 50
    show_full_result_count = False

    list_display = [
        'name', 'phone', 'product_display', 'region',
        'priority', 'status', 'amocrm_badge',
        'manager', 'created_at', 'action_buttons',
    ]
    list_editable = ['priority', 'status', 'manager']
    list_filter = ['status', 'priority', 'region']
    search_fields = ['name', 'phone', 'amocrm_lead_id']
    readonly_fields = ['created_at', 'amocrm_sent_at', 'amocrm_lead_link']
    autocomplete_fields = ['manager']
    actions = ['retry_failed_leads', 'export_to_excel']

    fieldsets = (
        ('Информация о клиенте', {
            'fields': ('name', 'phone', 'region', 'message', 'created_at'),
        }),
        ('Управление', {
            'fields': ('status', 'priority', 'manager', 'admin_comment'),
        }),
        ('amoCRM', {
            'fields': ('amocrm_status', 'amocrm_lead_link', 'amocrm_sent_at', 'amocrm_error'),
            'classes': ('collapse',),
        }),
    )

    class Media:
        css = {'all': ('css/amocrm_modal.css', 'css/contactform_admin.css')}
        js = ('js/amocrm_modal.js', 'js/contactform_admin.js')

    def product_display(self, obj):
        if not obj.product:
            return '—'
        return format_html('<span style="color:#1976d2;font-weight:600;">{}</span>', obj.product[:30])
    product_display.short_description = 'Модель'
    product_display.admin_order_field = 'product'

    def amocrm_badge(self, obj):
        if obj.amocrm_status == 'sent':
            return format_html(
                '<span style="background:#10b981;color:white;padding:5px 12px;'
                'border-radius:6px;font-weight:600;font-size:12px;">Отправлено</span>'
            )
        elif obj.amocrm_status == 'failed':
            error_text = (obj.amocrm_error or 'Неизвестная ошибка').replace('"', '&quot;').replace("'", '&#39;')
            return format_html(
                '<span class="amocrm-error-badge" data-error="{}" '
                'style="background:#ef4444;color:white;padding:5px 12px;border-radius:6px;'
                'font-weight:600;font-size:12px;cursor:pointer;">Ошибка</span>',
                error_text,
            )
        return format_html(
            '<span style="background:#f59e0b;color:white;padding:5px 12px;'
            'border-radius:6px;font-weight:600;font-size:12px;">Ожидает</span>'
        )
    amocrm_badge.short_description = 'amoCRM'
    amocrm_badge.admin_order_field = 'amocrm_status'

    def action_buttons(self, obj):
        view_url = (
            f"https://fawtrucks.amocrm.ru/leads/detail/{obj.amocrm_lead_id}"
            if obj.amocrm_lead_id
            else f"/admin/main/contactform/{obj.id}/change/"
        )
        view_title = 'Открыть в amoCRM' if obj.amocrm_lead_id else 'Просмотр заявки'
        return format_html(
            '<div style="display:flex;gap:8px;">'
            '<a href="{}" title="Редактировать"><img src="/static/media/icon-adminpanel/pencil.png" width="20" height="20"></a>'
            '<a href="{}" title="{}" target="_blank"><img src="/static/media/icon-adminpanel/eyes.png" width="20" height="20"></a>'
            '<a href="{}" title="Удалить" onclick="return confirm(\'Удалить заявку?\')"><img src="/static/media/icon-adminpanel/recycle-bin.png" width="20" height="20"></a>'
            '</div>',
            f'/admin/main/contactform/{obj.id}/change/',
            view_url, view_title,
            f'/admin/main/contactform/{obj.id}/delete/',
        )
    action_buttons.short_description = 'Действия'

    def amocrm_lead_link(self, obj):
        if obj.amocrm_lead_id:
            url = f"https://fawtrucks.amocrm.ru/leads/detail/{obj.amocrm_lead_id}"
            return format_html(
                '<a href="{}" target="_blank" style="color:#3b82f6;font-weight:600;">'
                'Открыть в amoCRM (ID: {})</a>',
                url, obj.amocrm_lead_id,
            )
        return '—'
    amocrm_lead_link.short_description = 'Ссылка на лид'

    @admin.action(description='Повторно отправить ошибочные заявки')
    def retry_failed_leads(self, request, queryset):
        failed_leads = queryset.filter(amocrm_status='failed')
        if not failed_leads.exists():
            self.message_user(request, 'Нет ошибочных заявок для повторной отправки.', level=messages.WARNING)
            return

        success_count = failed_count = 0
        for lead in failed_leads:
            try:
                lead.amocrm_status = 'pending'
                lead.amocrm_error = None
                lead.save()
                LeadSender.send_lead(lead)
                lead.refresh_from_db()
                if lead.amocrm_status == 'sent':
                    success_count += 1
                else:
                    failed_count += 1
            except Exception as exc:
                logger.error('Error retrying lead %s: %s', lead.id, exc)
                failed_count += 1

        if success_count:
            self.message_user(request, f'Успешно отправлено: {success_count}.', level=messages.SUCCESS)
        if failed_count:
            self.message_user(request, f'Ошибка отправки: {failed_count}.', level=messages.ERROR)

    @admin.action(description='Экспорт в Excel')
    def export_to_excel(self, request, queryset):
        try:
            if request.POST.get('select_across') == '1':
                queryset = self.get_queryset(request)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Заявки FAW UZ'

            headers = [
                'Номер', 'ФИО', 'Телефон', 'Модель', 'Регион', 'Сообщение',
                'Статус', 'Приоритет', 'Менеджер', 'Дата',
                'amoCRM Статус', 'amoCRM ID', 'amoCRM Дата', 'amoCRM Ошибка',
                'UTM Source', 'UTM Medium', 'UTM Campaign', 'UTM Term', 'UTM Content',
                'Referer', 'Visitor UID',
            ]
            ws.append(headers)

            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            utm_fill = PatternFill(start_color='1a6b3c', end_color='1a6b3c', fill_type='solid')

            for col_idx, cell in enumerate(ws[1], start=1):
                cell.fill = utm_fill if col_idx >= 15 else header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for idx, contact in enumerate(queryset, start=1):
                utm = {}
                if contact.utm_data:
                    try:
                        utm = json.loads(contact.utm_data)
                    except (json.JSONDecodeError, TypeError):
                        pass

                ws.append([
                    idx,
                    contact.name,
                    contact.phone,
                    contact.product[:30] if contact.product else '-',
                    contact.get_region_display(),
                    contact.message[:100] if contact.message else '-',
                    contact.get_status_display(),
                    contact.get_priority_display(),
                    contact.manager.username if contact.manager else '-',
                    contact.created_at.strftime('%d.%m.%Y %H:%M'),
                    contact.get_amocrm_status_display(),
                    contact.amocrm_lead_id or '-',
                    contact.amocrm_sent_at.strftime('%d.%m.%Y %H:%M') if contact.amocrm_sent_at else '-',
                    contact.amocrm_error[:100] if contact.amocrm_error else '-',
                    utm.get('utm_source', '-'),
                    utm.get('utm_medium', '-'),
                    utm.get('utm_campaign', '-'),
                    utm.get('utm_term', '-'),
                    utm.get('utm_content', '-'),
                    contact.referer[:100] if contact.referer else '-',
                    contact.visitor_uid or '-',
                ])

            for column in ws.columns:
                max_length = max(len(str(cell.value or '')) for cell in column)
                ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = (
                f'attachment; filename="faw_uz_contacts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            )
            wb.save(response)
            return response

        except Exception as exc:
            logger.error('Error exporting to Excel: %s', exc)
            self.message_user(request, f'Ошибка экспорта: {exc}', level=messages.ERROR)
            return redirect(request.path)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        formfield = super().formfield_for_foreignkey(db_field, request, **kwargs)
        if db_field.name == 'manager':
            for attr in ['can_add_related', 'can_change_related', 'can_delete_related', 'can_view_related']:
                setattr(formfield.widget, attr, False)
        return formfield

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if search_query := request.GET.get('q', '').strip():
            qs = qs.filter(
                Q(name__icontains=search_query) |
                Q(phone__icontains=search_query) |
                Q(amocrm_lead_id__icontains=search_query)
            )
        for field in ('status', 'amocrm_status', 'priority', 'region'):
            if val := request.GET.get(field, '').strip():
                qs = qs.filter(**{field: val})
        if product := request.GET.get('product', '').strip():
            qs = qs.filter(product__icontains=product)
        if date_from := request.GET.get('date_from', '').strip():
            try:
                qs = qs.filter(created_at__gte=timezone.make_aware(
                    datetime.strptime(date_from, '%Y-%m-%d').replace(hour=0, minute=0, second=0)
                ))
            except ValueError:
                pass
        if date_to := request.GET.get('date_to', '').strip():
            try:
                qs = qs.filter(created_at__lte=timezone.make_aware(
                    datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                ))
            except ValueError:
                pass
        return qs

    def get_changelist(self, request, **kwargs):
        from django.contrib.admin.views.main import ChangeList

        class CustomChangeList(ChangeList):
            def get_filters_params(self, params=None):
                lookup_params = super().get_filters_params(params)
                lookup_params.pop('date_from', None)
                lookup_params.pop('date_to', None)
                return lookup_params

        return CustomChangeList

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['regions'] = REGION_CHOICES
        extra_context['products'] = list(
            ContactForm.objects.exclude(product__isnull=True).exclude(product='')
            .values_list('product', flat=True).distinct().order_by('product')
        )
        return super().changelist_view(request, extra_context)

    def get_urls(self):
        urls = super().get_urls()
        return [
            path(
                '<int:object_id>/quick-update/',
                self.admin_site.admin_view(self.quick_update_view),
                name='contactform_quick_update',
            ),
        ] + urls

    def quick_update_view(self, request, object_id):
        if request.method != 'POST':
            return JsonResponse({'error': 'Method not allowed'}, status=405)
        try:
            obj = ContactForm.objects.get(pk=object_id)
            data = json.loads(request.body)
            if 'status' in data:
                obj.status = data['status']
            if 'priority' in data:
                obj.priority = data['priority']
            if 'manager' in data:
                obj.manager_id = data['manager'] if data['manager'] else None
            obj.save()
            return JsonResponse({'success': True})
        except ContactForm.DoesNotExist:
            return JsonResponse({'error': 'Object not found'}, status=404)
        except Exception as exc:
            return JsonResponse({'error': str(exc)}, status=500)


# ============ ВАКАНСИИ ============

class VacancyResponsibilityInline(TranslationStackedInline):
    model = VacancyResponsibility
    extra = 2
    fields = (('title', 'order'), 'text')


class VacancyRequirementInline(TranslationTabularInline):
    model = VacancyRequirement
    extra = 3
    fields = ('text', 'order')


class VacancyConditionInline(TranslationTabularInline):
    model = VacancyCondition
    extra = 3
    fields = ('text', 'order')


class VacancyIdealCandidateInline(TranslationTabularInline):
    model = VacancyIdealCandidate
    extra = 3
    fields = ('text', 'order')


@admin.register(Vacancy)
class VacancyAdmin(ContentAdminMixin, CustomReversionMixin, VersionAdmin, TabbedTranslationAdmin):
    list_display = ['title', 'is_active', 'applications_count', 'order', 'created_at']
    list_filter = ['is_active', 'created_at']
    search_fields = ['title', 'short_description']
    prepopulated_fields = {'slug': ('title',)}
    readonly_fields = ['created_at', 'updated_at', 'applications_count']
    inlines = [VacancyResponsibilityInline, VacancyRequirementInline, VacancyIdealCandidateInline, VacancyConditionInline]
    history_latest_first = True

    fieldsets = (
        ('Основная информация', {'fields': ('title', 'slug', 'short_description', 'is_active', 'order')}),
        ('Контакты', {'fields': ('contact_info',)}),
        ('Статистика', {'fields': ('applications_count', 'created_at', 'updated_at'), 'classes': ('collapse',)}),
    )

    def applications_count(self, obj):
        count = obj.get_applications_count()
        if count > 0:
            return format_html(
                '<a href="/admin/main/jobapplication/?vacancy__id__exact={}" '
                'style="color:#007bff;font-weight:bold;">{} Заявок</a>',
                obj.id, count,
            )
        return '0 заявок'
    applications_count.short_description = 'Заявки'

    def changelist_view(self, request, extra_context=None):
        return super().changelist_view(request, _add_recover_context(self, request, extra_context))


@admin.register(JobApplication)
class JobApplicationAdmin(LeadManagerMixin, admin.ModelAdmin):
    list_display = ['vacancy', 'region', 'applicant_name', 'resume_link', 'file_size_display', 'created_at', 'is_processed_badge']
    list_filter = ['is_processed', 'vacancy', 'region', 'created_at']
    search_fields = ['applicant_name', 'applicant_phone', 'applicant_email', 'vacancy__title']
    readonly_fields = ['created_at', 'file_size_display', 'resume_preview']
    date_hierarchy = 'created_at'
    autocomplete_fields = ['vacancy']

    fieldsets = (
        ('Информация', {'fields': ('vacancy', 'region', 'created_at')}),
        ('Резюме', {'fields': ('resume', 'file_size_display', 'resume_preview')}),
        ('Контакты', {'fields': ('applicant_name', 'applicant_phone', 'applicant_email')}),
        ('Обработка', {'fields': ('is_processed', 'admin_comment')}),
    )

    def resume_link(self, obj):
        if obj.resume:
            return format_html('<a href="{}" target="_blank" style="color:#007bff;font-weight:bold;">Скачать</a>', obj.resume.url)
        return '—'
    resume_link.short_description = 'Резюме'

    def file_size_display(self, obj):
        size = obj.get_file_size()
        return f'{size} MB' if size else '—'
    file_size_display.short_description = 'Размер'

    def resume_preview(self, obj):
        if obj.resume:
            if obj.resume.name.split('.')[-1].lower() in ['jpg', 'jpeg', 'png']:
                return format_html('<img src="{}" width="300" style="border-radius:8px;">', obj.resume.url)
            return format_html('<p style="color:#888;">{}</p>', obj.resume.name)
        return '—'
    resume_preview.short_description = 'Превью'

    def is_processed_badge(self, obj):
        if obj.is_processed:
            return format_html('<span style="color:green;font-weight:bold;">Рассмотрено</span>')
        return format_html('<span style="color:orange;font-weight:bold;">Новая</span>')
    is_processed_badge.short_description = 'Статус'


# ============ ИКОНКИ ============

@admin.register(FeatureIcon)
class FeatureIconAdmin(ContentAdminMixin, admin.ModelAdmin):
    list_display = ['icon_preview', 'name', 'order']
    list_editable = ['name', 'order']
    search_fields = ['name']

    def icon_preview(self, obj):
        if obj.icon:
            return format_html('<img src="{}" width="30" height="30"/>', obj.icon.url)
        return '—'
    icon_preview.short_description = 'Превью'


# ============ ДИЛЕРЫ ============

@admin.register(DealerService)
class DealerServiceAdmin(ContentAdminMixin, CustomReversionMixin, VersionAdmin, TabbedTranslationAdmin):
    list_display = ['name', 'slug', 'order', 'is_active', 'action_buttons']
    list_editable = ['order', 'is_active']
    search_fields = ['name']
    prepopulated_fields = {'slug': ('name',)}
    history_latest_first = True

    _BASE_SLUGS = frozenset(['sotuv', 'servis', 'ehtiyot-qismlar'])

    def has_delete_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        if obj and obj.slug in self._BASE_SLUGS:
            return False
        return super().has_delete_permission(request, obj)

    def action_buttons(self, obj):
        delete_btn = (
            '' if obj.slug in self._BASE_SLUGS
            else f'<a href="/admin/main/dealerservice/{obj.id}/delete/" onclick="return confirm(\'Удалить?\')">'
                 '<img src="/static/media/icon-adminpanel/recycle-bin.png" width="24" height="24"></a>'
        )
        return format_html(
            '<div style="display:flex;gap:8px;">'
            '<a href="{}"><img src="/static/media/icon-adminpanel/pencil.png" width="24" height="24"></a>'
            '<a href="/dealers/" target="_blank"><img src="/static/media/icon-adminpanel/eyes.png" width="24" height="24"></a>'
            '{}'
            '</div>',
            f'/admin/main/dealerservice/{obj.id}/change/',
            delete_btn,
        )
    action_buttons.short_description = 'Действия'

    def changelist_view(self, request, extra_context=None):
        return super().changelist_view(request, _add_recover_context(self, request, extra_context))


@admin.register(Dealer)
class DealerAdmin(ContentAdminMixin, CustomReversionMixin, VersionAdmin, TabbedTranslationAdmin):
    list_display = ['logo_preview', 'name', 'city', 'phone', 'services_list', 'is_active', 'order', 'action_buttons']
    list_filter = ['is_active', 'city', 'services']
    search_fields = ['name', 'city', 'address', 'manager']
    list_editable = ['is_active', 'order']
    readonly_fields = ['logo_preview', 'created_at', 'updated_at']
    history_latest_first = True

    fieldsets = (
        ('Основная информация', {'fields': ('name', 'city', 'address', 'logo', 'logo_preview')}),
        ('Координаты', {'fields': ('latitude', 'longitude')}),
        ('Контакты', {'fields': ('phone', 'email', 'website', 'manager')}),
        ('Рабочее время', {'fields': ('working_hours',)}),
        ('Услуги', {'fields': ('services',)}),
        ('Настройки', {'fields': ('is_active', 'order', 'created_at', 'updated_at'), 'classes': ('collapse',)}),
    )

    def formfield_for_manytomany(self, db_field, request, **kwargs):
        if db_field.name == 'services':
            kwargs['widget'] = forms.CheckboxSelectMultiple()
        return super().formfield_for_manytomany(db_field, request, **kwargs)

    def logo_preview(self, obj):
        if obj.logo:
            return format_html('<img src="{}" width="80" height="50" style="object-fit:contain;border-radius:4px;"/>', obj.logo.url)
        return '—'
    logo_preview.short_description = 'Логотип'

    def services_list(self, obj):
        services = obj.services.all()
        if services:
            tags = ' '.join(
                f'<span style="background:#e3f2fd;color:#1976d2;padding:4px 8px;border-radius:4px;font-size:11px;">{s.name}</span>'
                for s in services
            )
            return format_html(tags)
        return '—'
    services_list.short_description = 'Услуги'

    def action_buttons(self, obj):
        return format_html(
            '<div style="display:flex;gap:8px;">'
            '<a href="{}"><img src="/static/media/icon-adminpanel/pencil.png" width="24" height="24"></a>'
            '<a href="/dealers/" target="_blank"><img src="/static/media/icon-adminpanel/eyes.png" width="24" height="24"></a>'
            '<a href="{}" onclick="return confirm(\'Удалить?\')"><img src="/static/media/icon-adminpanel/recycle-bin.png" width="24" height="24"></a>'
            '</div>',
            f'/admin/main/dealer/{obj.id}/change/',
            f'/admin/main/dealer/{obj.id}/delete/',
        )
    action_buttons.short_description = 'Действия'

    def changelist_view(self, request, extra_context=None):
        return super().changelist_view(request, _add_recover_context(self, request, extra_context))


# ============ СТРАНИЦА "СТАТЬ ДИЛЕРОМ" ============

class DealerRequirementInline(TranslationTabularInline):
    model = DealerRequirement
    extra = 1
    fields = ('text', 'order')


@admin.register(BecomeADealerPage)
class BecomeADealerPageAdmin(ContentAdminMixin, TabbedTranslationAdmin):
    fieldsets = (
        ('Контент', {'fields': ('title', 'intro_text', 'subtitle', 'important_note')}),
        ('Контакты', {'fields': ('contact_phone', 'contact_email', 'contact_address')}),
    )
    inlines = [DealerRequirementInline]

    def has_add_permission(self, request):
        return not BecomeADealerPage.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False

    def changelist_view(self, request, extra_context=None):
        obj = BecomeADealerPage.get_instance()
        return self.changeform_view(request, str(obj.pk), '', extra_context)


# ============ ЗАЯВКИ НА ДИЛЕРСТВО ============

class BecomeADealerApplicationForm(forms.ModelForm):
    class Meta:
        model = BecomeADealerApplication
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if 'manager' in self.fields:
            for attr in ['can_add_related', 'can_change_related', 'can_delete_related', 'can_view_related']:
                setattr(self.fields['manager'].widget, attr, False)


@admin.register(BecomeADealerApplication)
class BecomeADealerApplicationAdmin(LeadManagerMixin, admin.ModelAdmin):
    form = BecomeADealerApplicationForm
    list_display = ['dealer_badge', 'name', 'company_name', 'phone', 'region', 'experience_years', 'status', 'priority', 'manager', 'created_at', 'action_buttons']
    search_fields = ['name', 'company_name', 'phone', 'message']
    list_editable = ['status', 'priority', 'manager']
    readonly_fields = ['created_at']
    autocomplete_fields = ['manager']
    date_hierarchy = 'created_at'
    actions = ['export_to_excel']

    fieldsets = (
        ('Заявитель', {'fields': ('name', 'company_name', 'experience_years', 'region', 'phone')}),
        ('Сообщение', {'fields': ('message',)}),
        ('Управление', {'fields': ('status', 'priority', 'manager', 'admin_comment', 'created_at')}),
    )

    def dealer_badge(self, obj):
        return format_html('<span style="background:#000;color:white;padding:4px 10px;border-radius:6px;font-size:11px;font-weight:600;">ДИЛЕРСТВО</span>')
    dealer_badge.short_description = 'Тип'

    def action_buttons(self, obj):
        return format_html(
            '<div style="display:flex;gap:8px;">'
            '<a href="{}"><img src="/static/media/icon-adminpanel/pencil.png" width="24" height="24"></a>'
            '<a href="/become-a-dealer/" target="_blank"><img src="/static/media/icon-adminpanel/eyes.png" width="24" height="24"></a>'
            '<a href="{}" onclick="return confirm(\'Удалить?\')"><img src="/static/media/icon-adminpanel/recycle-bin.png" width="24" height="24"></a>'
            '</div>',
            f'/admin/main/becomeadealerapplication/{obj.id}/change/',
            f'/admin/main/becomeadealerapplication/{obj.id}/delete/',
        )
    action_buttons.short_description = 'Действия'

    @admin.action(description='Экспорт в Excel')
    def export_to_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Заявки на дилерство'
        headers = ['№', 'ФИО', 'Компания', 'Опыт', 'Регион', 'Телефон', 'Статус', 'Приоритет', 'Менеджер', 'Дата']
        ws.append(headers)

        header_fill = PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for idx, app in enumerate(queryset, start=1):
            ws.append([
                idx, app.name, app.company_name or '-', app.experience_years or '-',
                app.get_region_display(), app.phone,
                app.get_status_display(), app.get_priority_display(),
                app.manager.username if app.manager else '-',
                app.created_at.strftime('%d.%m.%Y %H:%M'),
            ])

        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="dealer_applications_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response


# ============ ПРОДУКТЫ ============

class ProductCategoryFilter(admin.SimpleListFilter):
    title = 'категория'
    parameter_name = 'category_filter'

    def lookups(self, request, model_admin):
        return Product.CATEGORY_CHOICES

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(
                Q(category=self.value()) | Q(categories__contains=self.value())
            )
        return queryset


class PriceWidget(forms.MultiWidget):
    def __init__(self, attrs=None):
        widgets = [
            forms.CheckboxInput(attrs={'title': 'от / dan / from', 'style': 'width:auto;margin-right:6px;vertical-align:middle;'}),
            forms.NumberInput(attrs={'step': '1', 'style': 'width:220px;'}),
        ]
        super().__init__(widgets, attrs)

    def decompress(self, value):
        return [False, value]


class PriceField(forms.MultiValueField):
    widget = PriceWidget

    def __init__(self, *args, **kwargs):
        fields = [
            forms.BooleanField(required=False),
            forms.DecimalField(required=False, max_digits=15, decimal_places=2),
        ]
        kwargs.setdefault('require_all_fields', False)
        super().__init__(fields=fields, *args, **kwargs)

    def compress(self, data_list):
        return {
            'price_is_from': data_list[0] if data_list else False,
            'price': data_list[1] if len(data_list) > 1 else None,
        }


class ProductCategoriesForm(forms.ModelForm):
    selected_categories = forms.MultipleChoiceField(
        choices=Product.CATEGORY_CHOICES,
        widget=forms.CheckboxSelectMultiple,
        required=True,
        label='Категории',
        help_text='Выберите категории продукта',
    )
    price_combined = PriceField(
        label='Цена (UZS)',
        required=False,
        help_text='Перед ценой будет отображаться "от"',
    )

    class Meta:
        model = Product
        exclude = ['category', 'categories']

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance.pk:
            selected = []
            if self.instance.category:
                selected.append(self.instance.category)
            if self.instance.categories:
                selected.extend(c.strip() for c in self.instance.categories.split(',') if c.strip())
            self.fields['selected_categories'].initial = list(dict.fromkeys(selected))
            self.fields['price_combined'].initial = [self.instance.price_is_from, self.instance.price]

    def clean_selected_categories(self):
        categories = self.cleaned_data.get('selected_categories', [])
        if not categories:
            raise forms.ValidationError('Выберите хотя бы одну категорию')
        return categories

    def save(self, commit=True):
        instance = super().save(commit=False)
        selected = self.cleaned_data.get('selected_categories', [])
        if selected:
            instance.category = selected[0]
            instance.categories = ','.join(selected[1:]) if len(selected) > 1 else ''
        combined = self.cleaned_data.get('price_combined')
        if combined:
            instance.price_is_from = combined.get('price_is_from', False)
            instance.price = combined.get('price')
        if commit:
            instance.save()
        return instance


class ProductParameterInline(TranslationTabularInline):
    model = ProductParameter
    extra = 0
    fields = ('category', 'text', 'order')
    verbose_name = 'Параметр'
    verbose_name_plural = 'Параметры'

    class Media:
        js = ('js/admin/parameter_filter.js',)
        css = {'all': ('css/admin/parameter_filter.css',)}


class ProductFeatureInline(TranslationTabularInline):
    model = ProductFeature
    extra = 0
    max_num = 8
    fields = ('icon', 'name', 'order')
    verbose_name = 'Характеристика'
    verbose_name_plural = 'Характеристики с иконками'


class ProductCardSpecInline(TranslationTabularInline):
    model = ProductCardSpec
    extra = 0
    max_num = 4
    fields = ('icon', 'value', 'order')
    verbose_name = 'Спецификация'
    verbose_name_plural = 'Характеристики карточки'


class ProductGalleryInline(admin.TabularInline):
    model = ProductGallery
    extra = 1
    fields = ('image', 'order')
    verbose_name = 'Фото'
    verbose_name_plural = 'Галерея'


@admin.register(Product)
class ProductAdmin(ContentAdminMixin, CustomReversionMixin, VersionAdmin, TabbedTranslationAdmin):
    form = ProductCategoriesForm
    list_display = ['thumbnail', 'title', 'all_categories_display', 'is_active', 'is_featured', 'slider_order', 'order', 'seo_button']
    list_filter = [ProductCategoryFilter, 'is_active', 'is_featured']
    search_fields = ['title', 'slug']
    list_editable = ['is_active', 'is_featured', 'slider_order', 'order']
    prepopulated_fields = {'slug': ('title',)}
    history_latest_first = True
    actions = ['add_to_slider', 'remove_from_slider']
    list_per_page = 15
    show_full_result_count = False

    fieldsets = (
        ('Основная информация', {
            'fields': (
                ('title', 'slug'),
                'selected_categories',
                'price_combined',
                ('order', 'is_active', 'is_featured'),
                ('main_image', 'card_image'),
            ),
        }),
        ('Настройки главного слайдера', {
            'classes': ('collapse',),
            'fields': ('slider_image', 'slider_order', 'slider_price', ('slider_power', 'slider_fuel_consumption')),
        }),
    )
    inlines = [ProductParameterInline, ProductFeatureInline, ProductCardSpecInline, ProductGalleryInline]

    def thumbnail(self, obj):
        img = obj.card_image or obj.main_image
        if img:
            return format_html('<img src="{}" width="80" height="50" style="object-fit:cover;border-radius:4px;"/>', img.url)
        return '—'
    thumbnail.short_description = 'Фото'

    def all_categories_display(self, obj):
        categories = obj.get_all_categories()
        if not categories:
            return '—'
        name_map = dict(Product.CATEGORY_CHOICES)
        tags = []
        for idx, cat_slug in enumerate(categories):
            name = name_map.get(cat_slug, cat_slug)
            if idx == 0:
                tags.append(f'<span style="background:#1976d2;color:white;padding:4px 8px;border-radius:4px;font-size:11px;font-weight:600;">{name}</span>')
            else:
                tags.append(f'<span style="background:#d3ecff;color:#006ad3;padding:4px 8px;border-radius:4px;font-size:11px;">{name}</span>')
        return format_html(' '.join(tags))
    all_categories_display.short_description = 'Категории'

    def seo_button(self, obj):
        try:
            seo = PageMeta.objects.get(model='Product', key=str(obj.id))
            color = '#10b981' if seo.is_active else '#ef4444'
            status = 'Активно' if seo.is_active else 'Неактивно'
            return format_html(
                '<a href="/admin/main/pagemeta/{}/change/" title="{}" '
                'style="background:{};color:white;padding:6px 12px;border-radius:6px;'
                'text-decoration:none;font-weight:600;font-size:11px;">SEO</a>',
                seo.id, status, color,
            )
        except PageMeta.DoesNotExist:
            return format_html('<span style="color:#999;font-size:11px;">—</span>')
    seo_button.short_description = 'SEO'

    @admin.action(description='Добавить в слайдер')
    def add_to_slider(self, request, queryset):
        updated = queryset.update(is_featured=True)
        self.message_user(request, f'{updated} продуктов добавлено в слайдер.')

    @admin.action(description='Убрать из слайдера')
    def remove_from_slider(self, request, queryset):
        updated = queryset.update(is_featured=False)
        self.message_user(request, f'{updated} продуктов убрано из слайдера.')

    def changelist_view(self, request, extra_context=None):
        extra_context = _add_recover_context(self, request, extra_context)
        extra_context['featured_count'] = Product.objects.filter(is_featured=True, is_active=True).count()
        extra_context['show_slider_info'] = True
        return super().changelist_view(request, extra_context)

    def get_urls(self):
        return [
            path(
                'api/parameter-suggestions/',
                self.admin_site.admin_view(self.parameter_suggestions_api),
                name='parameter_suggestions_api',
            ),
        ] + super().get_urls()

    def parameter_suggestions_api(self, request):
        category = request.GET.get('category', '')
        if not category:
            return JsonResponse({'suggestions': []})
        suggestions = (
            ProductParameter.objects.filter(category=category)
            .values('text').annotate(usage_count=Count('id'))
            .order_by('-usage_count')[:20]
        )
        seen = set()
        result = []
        for item in suggestions:
            text = item['text']
            if text and text not in seen:
                seen.add(text)
                result.append({'text': text, 'count': item['usage_count']})
        return JsonResponse({'suggestions': result})


# ============ AMОCRM ТОКЕНЫ ============

@admin.register(AmoCRMToken)
class AmoCRMTokenAdmin(AmoCRMAdminMixin, admin.ModelAdmin):
    list_display = ['token_status', 'expires_display', 'time_left_display', 'action_buttons']

    def token_status(self, obj):
        if not obj.access_token:
            return format_html('<span style="background:#dc3545;color:white;padding:6px 12px;border-radius:6px;font-weight:600;">Не настроен</span>')
        if obj.is_expired():
            return format_html('<span style="background:#ffc107;color:#000;padding:6px 12px;border-radius:6px;font-weight:600;">Истекает скоро</span>')
        return format_html('<span style="background:#28a745;color:white;padding:6px 12px;border-radius:6px;font-weight:600;">Валиден</span>')
    token_status.short_description = 'Статус'

    def expires_display(self, obj):
        return obj.expires_at.strftime('%d.%m.%Y %H:%M') if obj.expires_at else '—'
    expires_display.short_description = 'Истекает'

    def time_left_display(self, obj):
        if not obj.expires_at:
            return '—'
        time_left = obj.expires_at - timezone.now()
        if time_left.total_seconds() < 0:
            return format_html('<span style="color:#dc3545;font-weight:600;">Истёк</span>')
        days, hours = time_left.days, int(time_left.seconds / 3600)
        minutes = int((time_left.seconds % 3600) / 60)
        color = '#dc3545' if time_left.total_seconds() < 3600 else '#ffc107' if time_left.total_seconds() < 7200 else '#28a745'
        if days > 0:
            text = f'{days} дн. {hours} ч.'
        elif hours > 0:
            text = f'{hours} ч. {minutes} мин.'
        else:
            text = f'{minutes} мин.'
        return format_html('<span style="color:{};font-weight:600;">{}</span>', color, text)
    time_left_display.short_description = 'Осталось'

    def action_buttons(self, obj):
        return format_html(
            '<div style="display:flex;gap:8px;flex-wrap:wrap;">'
            '<a href="/admin/main/amocrmtoken/refresh/" class="button" style="background:#007bff;color:white;padding:6px 12px;border-radius:4px;text-decoration:none;">Обновить токен</a>'
            '<a href="/admin/main/amocrmtoken/logs/" class="button" style="background:#dc3545;color:white;padding:6px 12px;border-radius:4px;text-decoration:none;">Логи ошибок</a>'
            '<a href="/admin/main/amocrmtoken/instructions/" class="button" style="background:#6c757d;color:white;padding:6px 12px;border-radius:4px;text-decoration:none;">Инструкция</a>'
            '</div>'
        )
    action_buttons.short_description = 'Действия'

    def get_urls(self):
        return [
            path('refresh/', self.admin_site.admin_view(self.refresh_token_view), name='amocrm_refresh'),
            path('logs/', self.admin_site.admin_view(self.logs_view), name='amocrm_logs'),
            path('instructions/', self.admin_site.admin_view(self.instructions_view), name='amocrm_instructions'),
        ] + super().get_urls()

    def refresh_token_view(self, request):
        try:
            token_obj = AmoCRMToken.get_instance()
            if not token_obj.refresh_token:
                messages.error(request, 'Refresh token не найден. Настройте токены заново.')
                return redirect('/admin/main/amocrmtoken/')
            TokenManager.refresh_token(token_obj)
            messages.success(request, f'Токен успешно обновлён. Истекает: {token_obj.expires_at.strftime("%d.%m.%Y %H:%M")}')
        except Exception as exc:
            messages.error(request, f'Ошибка обновления: {exc}')
        return redirect('/admin/main/amocrmtoken/')

    def logs_view(self, request):
        def _read_log(path, lines=50):
            if not os.path.exists(path):
                return []
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    return [l.strip() for l in f.readlines() if l.strip()][-lines:]
            except Exception as exc:
                return [f'Ошибка чтения: {exc}']

        context = {
            **self.admin_site.each_context(request),
            'title': 'Логи ошибок amoCRM',
            'amocrm_logs': _read_log(os.path.join(settings.BASE_DIR, 'logs', 'amocrm.log'), 50),
            'errors_logs': _read_log(os.path.join(settings.BASE_DIR, 'logs', 'errors.log'), 100),
        }
        return render(request, 'main/amocrm_logs.html', context)

    def instructions_view(self, request):
        token_obj = AmoCRMToken.get_instance()
        time_left_text = None
        if token_obj.expires_at:
            time_left = token_obj.expires_at - timezone.now()
            if time_left.total_seconds() < 0:
                time_left_text = 'Токен истёк'
            else:
                parts = []
                if time_left.days:
                    parts.append(f'{time_left.days} дн.')
                hours = int(time_left.seconds / 3600)
                if hours:
                    parts.append(f'{hours} ч.')
                minutes = int((time_left.seconds % 3600) / 60)
                if minutes:
                    parts.append(f'{minutes} мин.')
                time_left_text = ' '.join(parts) or 'Менее минуты'
        context = {
            **self.admin_site.each_context(request),
            'title': 'Инструкция: Обновление токенов amoCRM',
            'token': token_obj,
            'time_left_text': time_left_text,
        }
        return render(request, 'main/amocrm_instructions.html', context)


# ========== DASHBOARD ==========

@admin.register(Dashboard)
class DashboardAdmin(admin.ModelAdmin):
    change_list_template = 'main/dashboard/dashboard.html'

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return self.has_module_permission(request)

    def has_module_permission(self, request):
        if request.user.is_superuser:
            return True
        if request.user.groups.filter(name__in=['Главные админы', 'Лид-менеджеры']).exists():
            return True
        return request.user.has_perm('main.view_dashboard')

    def get_urls(self):
        return [
            path('api/data/', self.admin_site.admin_view(self.dashboard_api_data), name='dashboard_api_data'),
            path('api/products/', self.admin_site.admin_view(self.get_products_api), name='dashboard_api_products'),
            path('export/excel/', self.admin_site.admin_view(self.dashboard_export_excel), name='dashboard_export_excel'),
        ] + super().get_urls()

    def get_products_api(self, request):
        products = (
            ContactForm.objects.exclude(Q(product__isnull=True) | Q(product=''))
            .values_list('product', flat=True).distinct().order_by('product')
        )
        return JsonResponse({'success': True, 'products': list(products)})

    @method_decorator(never_cache)
    def changelist_view(self, request, extra_context=None):
        tz = timezone.get_current_timezone()
        today = timezone.now().astimezone(tz).strftime('%Y-%m-%d')
        context = {
            **self.admin_site.each_context(request),
            'title': 'Аналитика заявок VUM',
            'date_from': request.GET.get('date_from', today),
            'date_to': request.GET.get('date_to', today),
            'region': unquote(request.GET.get('region', '')),
            'product': unquote(request.GET.get('product', '')),
            'source': request.GET.get('source', ''),
            'regions': REGION_CHOICES,
            'products': [],
        }
        if extra_context:
            context.update(extra_context)
        return render(request, self.change_list_template, context)

    def _build_qs(self, request):
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        region = request.GET.get('region', '')
        product = request.GET.get('product', '')
        source = request.GET.get('source', '')

        tz = timezone.get_current_timezone()
        start_date = timezone.make_aware(datetime.strptime(date_from, '%Y-%m-%d'), tz)
        end_date = timezone.make_aware(
            datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59), tz
        )
        qs = ContactForm.objects.filter(created_at__gte=start_date, created_at__lte=end_date)
        if region:
            qs = qs.filter(region=region)
        if product:
            qs = qs.filter(product__icontains=product)
        if source:
            if source == 'direct':
                qs = qs.filter(Q(utm_data__isnull=True) | Q(utm_data=''))
            else:
                source_map = {
                    'google': 'google', 'yandex': 'yandex', 'instagram': 'ig',
                    'facebook': 'fb', 'telegram': 'telegram', 'tiktok': 'tiktok', 'youtube': 'youtube',
                }
                qs = qs.filter(utm_data__icontains=f'"utm_source":"{source_map.get(source, source)}"')
        return qs, start_date, end_date

    def dashboard_api_data(self, request):
        try:
            qs, start_date, end_date = self._build_qs(request)
            return JsonResponse({
                'success': True,
                'kpi': calculate_kpi(qs, start_date, end_date),
                'charts': get_chart_data(qs, start_date, end_date),
                'insights': generate_insights(qs, start_date, end_date),
            })
        except Exception as exc:
            logger.error('Dashboard API error: %s', exc)
            return JsonResponse({'success': False, 'error': str(exc)}, status=500)

    def dashboard_export_excel(self, request):
        try:
            qs, start_date, end_date = self._build_qs(request)
            date_from = request.GET.get('date_from')
            date_to = request.GET.get('date_to')
            kpi = calculate_kpi(qs, start_date, end_date)
            charts = get_chart_data(qs, start_date, end_date)

            wb = openpyxl.Workbook()
            ws_kpi = wb.active
            ws_kpi.title = 'KPI'
            ws_kpi['A1'] = 'Аналитика Dashboard FAW'
            ws_kpi['A1'].font = Font(bold=True, size=16)
            ws_kpi['A2'] = f'Период: {date_from} — {date_to}'
            for cell in ['A4', 'B4']:
                ws_kpi[cell].font = Font(bold=True)
            ws_kpi['A4'], ws_kpi['B4'] = 'Метрика', 'Значение'
            rows = [
                ('Всего заявок', kpi['total_leads']),
                ('Отправлено в amoCRM', kpi['amocrm_sent']),
                ('Конверсия amoCRM', f"{kpi['amocrm_conversion']}%"),
                ('Ср. время ответа', f"{kpi['avg_response_time']} мин"),
                ('Тренд', f"{kpi['trend']['value']}% ({kpi['trend']['direction']})"),
            ]
            for idx, (label, value) in enumerate(rows, start=5):
                ws_kpi[f'A{idx}'] = label
                ws_kpi[f'B{idx}'] = value

            def _add_sheet(name, headers, rows_data):
                ws = wb.create_sheet(name)
                ws.append(headers)
                for row in rows_data:
                    ws.append(row)
                return ws

            _add_sheet('Источники', ['Источник', 'Заявок', '%'], [
                [charts['sources']['labels'][i], charts['sources']['values'][i], f"{charts['sources']['percentages'][i]}%"]
                for i in range(len(charts['sources']['labels']))
            ])
            _add_sheet('Модели', ['Модель', 'Заявок', '%'], [
                [charts['top_models']['labels'][i], charts['top_models']['values'][i], f"{charts['top_models']['percentages'][i]}%"]
                for i in range(len(charts['top_models']['labels']))
            ])
            _add_sheet('Регионы', ['Регион', 'Заявок', '%'], [
                [charts['top_regions']['labels'][i], charts['top_regions']['values'][i], f"{charts['top_regions']['percentages'][i]}%"]
                for i in range(len(charts['top_regions']['labels']))
            ])
            _add_sheet('Временной анализ', ['Час', 'Заявок', '%', 'Топ модель'], [
                [item['hour'], item['count'], f"{item['percent']}%", item['top_model']]
                for item in charts['time_analysis']['by_hours']
            ])
            _add_sheet('Повторные клиенты', ['Имя', 'Телефон', 'Заявок', 'Модели', 'Интервал (дней)', 'Последняя заявка'], [
                [c['name'], c['phone'], c['count'], c['models'], c['interval_days'], c['last_date']]
                for c in charts['behavior']['clients_list']
            ])

            header_style = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            for ws in wb:
                for cell in ws[1]:
                    cell.font = header_style
                    cell.fill = header_fill

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="dashboard_faw_{date_from}_{date_to}.xlsx"'
            wb.save(response)
            return response
        except Exception as exc:
            return HttpResponse(f'Ошибка экспорта: {exc}', content_type='text/plain', status=500)


# ============ АКЦИИ ============

@admin.register(Promotion)
class PromotionAdmin(ContentAdminMixin, CustomReversionMixin, VersionAdmin, TabbedTranslationAdmin):
    list_display = ['image_preview', 'title', 'is_active', 'show_on_homepage', 'priority', 'start_date', 'end_date', 'status_badge', 'action_buttons']
    list_editable = ['is_active', 'show_on_homepage', 'priority']
    list_filter = ['is_active', 'show_on_homepage', 'start_date', 'end_date']
    search_fields = ['title', 'description']
    date_hierarchy = 'start_date'
    history_latest_first = True
    readonly_fields = ['image_preview_large', 'created_at', 'updated_at']

    fieldsets = (
        (_('Основная информация'), {'fields': ('title', 'description', 'image', 'image_preview_large', 'link', 'button_text')}),
        (_('Настройки отображения'), {'fields': ('is_active', 'show_on_homepage', 'priority')}),
        (_('Период действия'), {'fields': ('start_date', 'end_date')}),
        (_('Техническая информация'), {'fields': ('created_at', 'updated_at'), 'classes': ('collapse',)}),
    )

    def status_badge(self, obj):
        if obj.is_valid():
            return format_html('<span style="padding:3px 10px;background:#28a745;color:white;border-radius:3px;">Активна</span>')
        return format_html('<span style="padding:3px 10px;background:#dc3545;color:white;border-radius:3px;">Неактивна</span>')
    status_badge.short_description = 'Статус'

    def image_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" style="width:50px;height:50px;object-fit:cover;border-radius:4px;"/>', obj.image.url)
        return '—'
    image_preview.short_description = 'Превью'

    def image_preview_large(self, obj):
        if obj.image:
            return format_html('<img src="{}" style="max-width:400px;max-height:300px;object-fit:contain;border-radius:8px;border:1px solid #ddd;"/>', obj.image.url)
        return '—'
    image_preview_large.short_description = 'Изображение'

    def action_buttons(self, obj):
        return format_html(
            '<div style="display:flex;gap:8px;">'
            '<a href="{}"><img src="/static/media/icon-adminpanel/pencil.png" width="24" height="24"></a>'
            '<a href="/" target="_blank"><img src="/static/media/icon-adminpanel/eyes.png" width="24" height="24"></a>'
            '<a href="{}" onclick="return confirm(\'Удалить акцию?\')"><img src="/static/media/icon-adminpanel/recycle-bin.png" width="24" height="24"></a>'
            '</div>',
            f'/admin/main/promotion/{obj.id}/change/',
            f'/admin/main/promotion/{obj.id}/delete/',
        )
    action_buttons.short_description = 'Действия'

    def changelist_view(self, request, extra_context=None):
        return super().changelist_view(request, _add_recover_context(self, request, extra_context))

    class Media:
        js = (
            'http://ajax.googleapis.com/ajax/libs/jquery/1.9.1/jquery.min.js',
            'http://ajax.googleapis.com/ajax/libs/jqueryui/1.10.2/jquery-ui.min.js',
            'modeltranslation/js/tabbed_translation_fields.js',
        )
        css = {'screen': ('modeltranslation/css/tabbed_translation_fields.css',)}


# ============ SEO META DATA ============

@admin.register(PageMeta)
class PageMetaAdmin(ContentAdminMixin, TabbedTranslationAdmin):
    form = PageMetaAdminForm
    list_display = ['status_compact', 'model_badge', 'content_with_link', 'title_preview', 'content_created_at', 'updated_at', 'action_buttons']
    list_filter = ['model', 'is_active']
    search_fields = ['key', 'title', 'title_uz', 'title_ru', 'title_en']
    readonly_fields = ['created_at', 'updated_at', 'og_image_preview_large']

    fieldsets = (
        ('Идентификация страницы', {'fields': ('model', 'key', 'is_active')}),
        ('Базовые META теги', {'fields': ('title', 'description', 'keywords')}),
        ('Open Graph теги', {
            'fields': ('og_title', 'og_description', 'og_url', 'og_type', 'og_site_name', 'og_image', 'og_image_preview_large'),
            'classes': ('collapse',),
        }),
        ('Служебная информация', {'fields': ('created_at', 'updated_at'), 'classes': ('collapse',)}),
    )

    class Media:
        css = {'all': ('css/admin_seo_help.css',)}
        js = ('js/admin_seo_help.js',)

    def get_queryset(self, request):
        qs = super().get_queryset(request).order_by()
        if model_filter := request.GET.get('model'):
            qs = qs.filter(model=model_filter)
        return qs.annotate(
            priority=Case(When(is_active=True, then=0), When(is_active=False, then=1), output_field=IntegerField())
        ).order_by('priority', 'model', '-updated_at')

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        stats = {s['model']: s['count'] for s in PageMeta.objects.values('model').annotate(count=Count('id'))}
        total = sum(stats.values())
        current_model = request.GET.get('model', 'all')
        extra_context['tabs'] = [
            {'label': f'Все ({total})', 'url': '?', 'active': current_model == 'all'},
            {'label': f'Статические ({stats.get("Page", 0)})', 'url': '?model=Page', 'active': current_model == 'Page'},
            {'label': f'Новости ({stats.get("Post", 0)})', 'url': '?model=Post', 'active': current_model == 'Post'},
            {'label': f'Продукты ({stats.get("Product", 0)})', 'url': '?model=Product', 'active': current_model == 'Product'},
        ]
        return super().changelist_view(request, extra_context)

    def status_compact(self, obj):
        obj_id = f'#{obj.key} ' if obj.model in ['Post', 'Product'] else ''
        if obj.is_active:
            return format_html('<span style="color:#10b981;font-weight:600;">{}Готово</span>', obj_id)
        elif not obj.title:
            return format_html('<span style="color:#ef4444;font-weight:600;">{}Требует заполнения</span>', obj_id)
        return format_html('<span style="color:#f59e0b;font-weight:600;">{}В работе</span>', obj_id)
    status_compact.short_description = 'Статус'
    status_compact.admin_order_field = 'is_active'

    def model_badge(self, obj):
        colors = {'Page': '#3b82f6', 'Post': '#10b981', 'Product': '#f59e0b'}
        return format_html(
            '<span style="background:{};color:white;padding:4px 10px;border-radius:6px;font-size:12px;font-weight:600;">{}</span>',
            colors.get(obj.model, '#6b7280'), obj.get_model_display(),
        )
    model_badge.short_description = 'Тип'

    def content_with_link(self, obj):
        _PAGE_NAMES = {
            'home': 'Главная страница', 'about': 'О нас', 'contact': 'Контакты',
            'services': 'Сервис', 'lizing': 'Лизинг', 'become-a-dealer': 'Стать дилером',
            'jobs': 'Вакансии', 'news': 'Новости (список)', 'dealers': 'Дилеры (список)',
        }
        if obj.model == 'Post':
            try:
                news = News.objects.get(id=int(obj.key))
                title = news.title[:60] + ('...' if len(news.title) > 60 else '')
                return format_html(
                    '<div style="max-width:350px;"><strong style="display:block;margin-bottom:4px;">{}</strong>'
                    '<a href="/admin/main/news/{}/change/" style="font-size:14px;color:#3b82f6;">Открыть новость</a></div>',
                    title, news.id,
                )
            except (News.DoesNotExist, ValueError):
                return format_html('<span style="color:#999;">Новость не найдена</span>')
        elif obj.model == 'Product':
            try:
                product = Product.objects.get(id=int(obj.key))
                title = product.title[:60] + ('...' if len(product.title) > 60 else '')
                return format_html(
                    '<div style="max-width:350px;"><strong style="display:block;margin-bottom:4px;">{}</strong>'
                    '<a href="/admin/main/product/{}/change/" style="font-size:14px;color:#3b82f6;">Открыть продукт</a></div>',
                    title, product.id,
                )
            except (Product.DoesNotExist, ValueError):
                return format_html('<span style="color:#999;">Продукт не найден</span>')
        elif obj.model == 'Page':
            return format_html('<div style="max-width:350px;color:#666;"><em>{}</em></div>', _PAGE_NAMES.get(obj.key, obj.key))
        return '—'
    content_with_link.short_description = 'Контент'

    def content_created_at(self, obj):
        model_map = {'Post': News, 'Product': Product}
        model_cls = model_map.get(obj.model)
        if model_cls:
            try:
                instance = model_cls.objects.get(id=int(obj.key))
                return instance.created_at.strftime('%d.%m.%Y')
            except (model_cls.DoesNotExist, ValueError):
                pass
        return '—'
    content_created_at.short_description = 'Создан'

    def title_preview(self, obj):
        if not obj.title:
            return format_html('<span style="color:#ef4444;font-style:italic;">(не заполнено)</span>')
        title = obj.title[:50] + ('...' if len(obj.title) > 50 else '')
        return format_html('<span style="color:#374151;">{}</span>', title)
    title_preview.short_description = 'SEO Title'

    def action_buttons(self, obj):
        return format_html(
            '<div style="display:flex;gap:8px;">'
            '<a href="{}"><img src="/static/media/icon-adminpanel/pencil.png" width="24" height="24"></a>'
            '<a href="{}" target="_blank"><img src="/static/media/icon-adminpanel/eyes.png" width="24" height="24"></a>'
            '<a href="{}" onclick="return confirm(\'Удалить мета-данные?\')"><img src="/static/media/icon-adminpanel/recycle-bin.png" width="24" height="24"></a>'
            '</div>',
            f'/admin/main/pagemeta/{obj.id}/change/',
            obj.get_full_url(),
            f'/admin/main/pagemeta/{obj.id}/delete/',
        )
    action_buttons.short_description = 'Действия'

    def og_image_preview_large(self, obj):
        if obj.og_image:
            return format_html('<img src="{}" style="max-width:400px;max-height:200px;object-fit:contain;border-radius:8px;border:1px solid #ddd;"/>', obj.og_image.url)
        return '—'
    og_image_preview_large.short_description = 'Превью изображения'

    def save_model(self, request, obj, form, change):
        if not obj.og_title:
            obj.og_title = obj.title
        if not obj.og_description:
            obj.og_description = obj.description
        if not obj.og_url:
            obj.og_url = obj.get_full_url()
        super().save_model(request, obj, form, change)


# ========== FAQ ==========

@admin.register(FAQItem)
class FAQItemAdmin(TabbedTranslationAdmin):
    list_display = ('question', 'order', 'is_active')
    list_editable = ('order', 'is_active')
    list_display_links = ('question',)
    ordering = ('order',)


# ========== TELEGRAM BOT ADMIN ==========

@admin.register(TelegramUser)
class TelegramUserAdmin(admin.ModelAdmin):
    list_display = ['telegram_id', 'full_name', 'phone', 'region', 'language', 'status', 'total_requests', 'is_blocked', 'last_active']
    list_filter = ['language', 'status', 'region', 'is_blocked', 'notifications_enabled']
    search_fields = ['telegram_id', 'username', 'first_name', 'last_name', 'phone']
    readonly_fields = ['telegram_id', 'last_active', 'created_at', 'total_requests']
    list_editable = ['status']
    ordering = ['-last_active']

    fieldsets = (
        ('Telegram', {'fields': ('telegram_id', 'username', 'language', 'is_blocked', 'notifications_enabled')}),
        ('Личные данные', {'fields': ('last_name', 'first_name', 'middle_name', 'phone', 'age', 'region')}),
        ('Статус и активность', {'fields': ('status', 'total_requests', 'last_active', 'created_at')}),
        ('Реферальная система', {'fields': ('referral_code', 'referred_by'), 'classes': ('collapse',)}),
    )


@admin.register(TestDriveRequest)
class TestDriveRequestAdmin(admin.ModelAdmin):
    list_display = ['client_name', 'client_phone', 'product', 'dealer', 'preferred_date', 'preferred_time', 'status', 'reminder_sent', 'created_at']
    list_filter = ['status', 'dealer', 'reminder_sent', 'feedback_requested']
    search_fields = ['client_name', 'client_phone']
    list_editable = ['status']
    readonly_fields = ['created_at', 'updated_at']
    date_hierarchy = 'preferred_date'

    fieldsets = (
        ('Клиент', {'fields': ('user', 'client_name', 'client_phone')}),
        ('Заявка', {'fields': ('dealer', 'product', 'preferred_date', 'preferred_time', 'status')}),
        ('Уведомления', {'fields': ('reminder_sent', 'feedback_requested')}),
        ('Служебное', {'fields': ('admin_comment', 'created_at', 'updated_at'), 'classes': ('collapse',)}),
    )


@admin.register(BotMessage)
class BotMessageAdmin(admin.ModelAdmin):
    list_display = ['key', 'language', 'description', 'updated_at']
    list_filter = ['language']
    search_fields = ['key', 'text', 'description']
    ordering = ['key', 'language']
    list_per_page = 50
    readonly_fields = ('updated_at',)

    fieldsets = (
        (None, {'fields': ('key', 'language', 'text', 'description')}),
        ('Служебное', {'fields': ('updated_at',), 'classes': ('collapse',)}),
    )


@admin.register(BotMenuItem)
class BotMenuItemAdmin(admin.ModelAdmin):
    list_display = ['key', 'order', 'emoji', 'label_ru', 'label_uz', 'label_en', 'is_active']
    list_display_links = ['key']
    list_editable = ['order', 'is_active']
    list_filter = ['is_active']
    ordering = ['order']


@admin.register(BotBroadcast)
class BotBroadcastAdmin(admin.ModelAdmin):
    list_display = ['title', 'target', 'status', 'scheduled_at', 'sent_at', 'total_recipients', 'sent_count']
    list_filter = ['status', 'target']
    search_fields = ['title']
    ordering = ['-created_at']
    actions = ['send_now']

    fieldsets = (
        ('Контент', {'fields': ('title', 'text_ru', 'text_uz', 'text_en', 'image', 'button_text', 'button_url')}),
        ('Аудитория', {'fields': ('target', 'target_region', 'scheduled_at')}),
        ('Статус', {'fields': ('status',)}),
        ('Статистика', {
            'fields': ('total_recipients', 'sent_count', 'failed_count', 'blocked_count', 'sent_at'),
            'classes': ('collapse',),
        }),
        ('Служебное', {'fields': ('created_by', 'created_at'), 'classes': ('collapse',)}),
    )
    readonly_fields = ['total_recipients', 'sent_count', 'failed_count', 'blocked_count', 'sent_at', 'created_at']

    @admin.action(description='Запустить рассылку сейчас')
    def send_now(self, request, queryset):
        broadcasts = list(queryset.filter(status__in=['draft', 'scheduled']))
        if not broadcasts:
            self.message_user(request, 'Нет рассылок для отправки (только draft/scheduled).', level='warning')
            return

        config = BotConfig.get_instance()
        if not config or not config.bot_token:
            self.message_user(request, 'BotConfig не настроен — токен отсутствует.', level='error')
            return

        async def _run():
            bot = Bot(token=config.bot_token, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
            try:
                for broadcast in broadcasts:
                    qs = TelegramUser.objects.filter(is_blocked=False)
                    target_filters = {
                        'ru': {'language': 'ru'}, 'uz': {'language': 'uz'}, 'en': {'language': 'en'},
                        'hot': {'status': 'hot'}, 'vip': {'status': 'vip'},
                    }
                    if broadcast.target in target_filters:
                        qs = qs.filter(**target_filters[broadcast.target])
                    elif broadcast.target == 'active_30':
                        qs = qs.filter(last_active__gte=timezone.now() - timedelta(days=30))
                    elif broadcast.target == 'region' and broadcast.target_region:
                        qs = qs.filter(region=broadcast.target_region)

                    recipients = list(qs.values('telegram_id', 'language'))
                    BotBroadcast.objects.filter(pk=broadcast.pk).update(
                        status='sending', total_recipients=len(recipients)
                    )

                    sent = failed = blocked = 0
                    for recipient in recipients:
                        telegram_id = recipient['telegram_id']
                        lang = recipient['language'] or 'ru'
                        text = broadcast.get_text(lang)
                        if not text:
                            failed += 1
                            continue

                        reply_markup = None
                        if broadcast.button_text and broadcast.button_url:
                            reply_markup = InlineKeyboardMarkup(inline_keyboard=[[
                                InlineKeyboardButton(text=broadcast.button_text, url=broadcast.button_url)
                            ]])
                        try:
                            if broadcast.image:
                                await bot.send_photo(chat_id=telegram_id, photo=broadcast.image.url, caption=text, reply_markup=reply_markup)
                            else:
                                await bot.send_message(chat_id=telegram_id, text=text, reply_markup=reply_markup)
                            sent += 1
                        except TelegramForbiddenError:
                            blocked += 1
                            TelegramUser.objects.filter(telegram_id=telegram_id).update(is_blocked=True)
                        except Exception as exc:
                            logger.error('Broadcast error telegram_id=%s: %s', telegram_id, exc)
                            failed += 1
                        await asyncio.sleep(0.05)

                    BotBroadcast.objects.filter(pk=broadcast.pk).update(
                        status='done', sent_count=sent, failed_count=failed,
                        blocked_count=blocked, sent_at=timezone.now(),
                    )
            finally:
                await bot.session.close()

        try:
            asyncio.run(_run())
            self.message_user(request, f'Рассылка завершена. Обработано: {len(broadcasts)}.')
        except Exception as exc:
            logger.error('send_now admin action failed: %s', exc)
            self.message_user(request, f'Ошибка рассылки: {exc}', level='error')


@admin.register(BotContacts)
class BotContactsAdmin(admin.ModelAdmin):
    fieldsets = (
        ('Основные контакты', {'fields': ('phone_main', 'phone_secondary', 'email')}),
        ('Адрес', {'fields': ('address_ru', 'address_uz', 'address_en')}),
        ('Рабочее время', {'fields': ('working_hours_ru', 'working_hours_uz', 'working_hours_en')}),
        ('Соцсети', {'fields': ('telegram_channel', 'instagram', 'youtube', 'facebook', 'linkedin')}),
        ('Сайт и карта', {'fields': ('website', 'map_url')}),
        ('Служебное', {'fields': ('updated_at',), 'classes': ('collapse',)}),
    )
    readonly_fields = ('updated_at',)

    def has_add_permission(self, request):
        return not BotContacts.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(BotConfig)
class BotConfigAdmin(admin.ModelAdmin):
    readonly_fields = ('updated_at', 'bot_status_display', 'bot_control_buttons')

    fieldsets = (
        ('Основные настройки', {'fields': ('bot_token', 'bot_username', 'is_active')}),
        ('Управление ботом', {
            'fields': ('bot_status_display', 'bot_control_buttons'),
            'description': 'Управление процессом бота на сервере. Требует настроенного systemd сервиса.',
        }),
        ('Уведомления', {'fields': ('notify_chat_id', 'notify_chat_id_2')}),
        ('Сайт и режим работы', {'fields': ('site_url', 'use_webhook', 'webhook_url')}),
        ('Рабочие часы', {'fields': ('work_hours_start', 'work_hours_end')}),
        ('Файлы', {'fields': ('catalog_file',)}),
        ('Служебное', {'fields': ('updated_at',), 'classes': ('collapse',)}),
    )

    def _get_bot_status(self) -> tuple[bool, str]:
        try:
            result = subprocess.run(
                ['systemctl', 'is-active', BOT_SERVICE_NAME],
                capture_output=True, text=True, timeout=5,
            )
            is_running = result.stdout.strip() == 'active'
            return is_running, result.stdout.strip()
        except FileNotFoundError:
            return False, 'systemd недоступен (локально)'
        except subprocess.TimeoutExpired:
            return False, 'timeout'
        except Exception as exc:
            logger.error('Bot status check failed: %s', exc)
            return False, f'Ошибка: {exc}'

    def bot_status_display(self, obj):
        is_running, status_text = self._get_bot_status()
        if 'недоступен' in status_text:
            color, label = '#888', 'Локальная разработка — статус недоступен'
        elif is_running:
            color, label = '#28a745', 'Работает'
        else:
            color, label = '#dc3545', f'Остановлен ({status_text})'
        return format_html('<span style="color:{};font-weight:bold;">{}</span>', color, label)
    bot_status_display.short_description = 'Статус сервиса'

    def bot_control_buttons(self, obj):
        return format_html(
            '<a class="button" href="bot-start/" style="background:#28a745;color:white;padding:6px 12px;border-radius:4px;text-decoration:none;margin-right:8px;">Запустить</a>'
            '<a class="button" href="bot-stop/" style="background:#dc3545;color:white;padding:6px 12px;border-radius:4px;text-decoration:none;margin-right:8px;">Остановить</a>'
            '<a class="button" href="bot-restart/" style="background:#fd7e14;color:white;padding:6px 12px;border-radius:4px;text-decoration:none;">Перезапустить</a>'
        )
    bot_control_buttons.short_description = 'Управление'

    def get_urls(self):
        return [
            path('<path:object_id>/bot-start/', self.admin_site.admin_view(self._bot_start), name='botconfig-start'),
            path('<path:object_id>/bot-stop/', self.admin_site.admin_view(self._bot_stop), name='botconfig-stop'),
            path('<path:object_id>/bot-restart/', self.admin_site.admin_view(self._bot_restart), name='botconfig-restart'),
        ] + super().get_urls()

    def _run_systemctl(self, request, action: str):
        try:
            result = subprocess.run(
                ['sudo', 'systemctl', action, BOT_SERVICE_NAME],
                capture_output=True, text=True, timeout=15,
            )
            if result.returncode == 0:
                logger.info('Bot service %s: success admin=%s', action, request.user)
                return True
            logger.error('Bot service %s failed: %s', action, result.stderr)
            return False
        except FileNotFoundError:
            return None
        except subprocess.TimeoutExpired:
            return None
        except Exception as exc:
            logger.error('Bot control error action=%s: %s', action, exc)
            return False

    def _bot_action(self, request, action, success_msg, fail_msg):
        result = self._run_systemctl(request, action)
        if result is True:
            self.message_user(request, success_msg, messages.SUCCESS)
        elif result is None:
            self.message_user(request, 'systemd недоступен. Используйте терминал.', messages.WARNING)
        else:
            self.message_user(request, fail_msg, messages.ERROR)
        return HttpResponseRedirect('../')

    def _bot_start(self, request, object_id):
        return self._bot_action(request, 'start', 'Бот запущен.', 'Ошибка запуска. Проверьте логи.')

    def _bot_stop(self, request, object_id):
        return self._bot_action(request, 'stop', 'Бот остановлен.', 'Ошибка остановки. Проверьте логи.')

    def _bot_restart(self, request, object_id):
        return self._bot_action(request, 'restart', 'Бот перезапущен.', 'Ошибка перезапуска. Проверьте логи.')


# ========== КОМАНДА ==========

class TeamMemberInline(ContentAdminMixin, TranslationTabularInline):
    model = TeamMember
    extra = 1
    fields = (
        'order', 'is_active', 'photo',
        'name_uz', 'name_ru', 'name_en',
        'position_uz', 'position_ru', 'position_en',
    )
    ordering = ('order',)
    show_change_link = True


@admin.register(TeamDepartment)
class TeamDepartmentAdmin(ContentAdminMixin, TabbedTranslationAdmin):
    list_display = ('order', 'title', 'members_count', 'is_active')
    list_display_links = ('title',)
    list_editable = ('order', 'is_active')
    list_filter = ('is_active',)
    search_fields = ('title',)
    ordering = ('order',)
    inlines = [TeamMemberInline]

    fieldsets = (
        ('Основное', {
            'fields': ('title', 'order', 'is_active'),
        }),
    )

    def members_count(self, obj):
        count = obj.members.filter(is_active=True).count()
        return format_html('<b style="color:#28a745">{}</b>', count)
    members_count.short_description = 'Сотрудников'


@admin.register(TeamMember)
class TeamMemberAdmin(ContentAdminMixin, TabbedTranslationAdmin):
    list_display = ('order', 'photo_preview', 'name', 'position', 'department', 'is_active')
    list_display_links = ('name',)
    list_editable = ('order', 'is_active')
    list_filter = ('department', 'is_active')
    search_fields = ('name', 'position', 'department__title')
    ordering = ('department__order', 'order')

    fieldsets = (
        ('Основное', {
            'fields': ('department', 'order', 'is_active'),
        }),
        ('Данные сотрудника', {
            'fields': ('photo', 'name', 'position'),
        }),
    )

    def photo_preview(self, obj):
        if obj.photo:
            return format_html(
                '<img src="{}" style="width:48px;height:48px;object-fit:cover;border-radius:50%;">',
                obj.photo.url,
            )
        return '—'
    photo_preview.short_description = 'Фото'